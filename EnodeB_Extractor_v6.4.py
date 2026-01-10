import sys
import os
import pandas as pd
from PyQt6.QtWidgets import QApplication, QWidget, QPushButton, QFileDialog, QVBoxLayout, QLabel, QLineEdit, \
    QHBoxLayout, QMessageBox, QFrame
from PyQt6.QtGui import QFontMetrics, QIcon, QMovie, QPixmap, QFont
from PyQt6.QtCore import Qt, QTimer, QSize, QDateTime, QElapsedTimer, QThread, pyqtSignal
import subprocess
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import winsound
import traceback
import psutil
from collections import OrderedDict
import pickle
import hashlib
import json
from datetime import datetime, timedelta


class DatabaseCache:
    """Handles caching of loaded Excel/SQL databases to reduce loading time"""
    
    def __init__(self, cache_dir=None):
        # Use temp directory if not specified
        if cache_dir is None:
            cache_dir = os.path.join(os.path.expanduser('~'), '.enodeb_extractor_cache')
        
        self.cache_dir = cache_dir
        os.makedirs(self.cache_dir, exist_ok=True)
        
        self.cache_validity_hours = 24  # Cache valid for 24 hours (daily database update)
    
    def _get_file_signature(self, file_path):
        """Generate a signature for the file based on path, size, and modification time"""
        try:
            stat = os.stat(file_path)
            signature = {
                'path': file_path,
                'size': stat.st_size,
                'mtime': stat.st_mtime
            }
            return signature
        except Exception:
            return None
    
    def _get_cache_path(self, file_path):
        """Generate cache file path based on the source file"""
        # Create a hash of the file path to use as cache filename
        file_hash = hashlib.md5(file_path.encode()).hexdigest()
        cache_file = os.path.join(self.cache_dir, f"cache_{file_hash}.pkl")
        metadata_file = os.path.join(self.cache_dir, f"cache_{file_hash}.json")
        return cache_file, metadata_file
    
    def is_cache_valid(self, file_path):
        """Check if cached data exists and is still valid"""
        cache_file, metadata_file = self._get_cache_path(file_path)
        
        # Check if cache files exist
        if not os.path.exists(cache_file) or not os.path.exists(metadata_file):
            return False
        
        try:
            # Load metadata
            with open(metadata_file, 'r') as f:
                metadata = json.load(f)
            
            # Check if source file has changed
            current_signature = self._get_file_signature(file_path)
            if current_signature is None:
                return False
            
            cached_signature = metadata.get('file_signature', {})
            
            # Validate file hasn't changed
            if (current_signature['path'] != cached_signature.get('path') or
                current_signature['size'] != cached_signature.get('size') or
                current_signature['mtime'] != cached_signature.get('mtime')):
                return False
            
            # Check cache age (24 hours validity)
            cache_time = datetime.fromisoformat(metadata.get('cache_time'))
            cache_age = datetime.now() - cache_time
            
            if cache_age > timedelta(hours=self.cache_validity_hours):
                return False
            
            return True
            
        except Exception:
            return False
    
    def load_from_cache(self, file_path):
        """Load dataframe from cache"""
        cache_file, metadata_file = self._get_cache_path(file_path)
        
        try:
            with open(cache_file, 'rb') as f:
                df = pickle.load(f)
            return df
        except Exception:
            return None
    
    def save_to_cache(self, file_path, df):
        """Save dataframe to cache with metadata"""
        cache_file, metadata_file = self._get_cache_path(file_path)
        
        try:
            # Save dataframe as pickle
            with open(cache_file, 'wb') as f:
                pickle.dump(df, f, protocol=pickle.HIGHEST_PROTOCOL)
            
            # Save metadata
            metadata = {
                'file_signature': self._get_file_signature(file_path),
                'cache_time': datetime.now().isoformat(),
                'rows': len(df),
                'columns': len(df.columns)
            }
            
            with open(metadata_file, 'w') as f:
                json.dump(metadata, f, indent=2)
            
            return True
        except Exception:
            return False
    
    def get_cache_info(self, file_path):
        """Get information about cached data"""
        cache_file, metadata_file = self._get_cache_path(file_path)
        
        if not os.path.exists(metadata_file):
            return None
        
        try:
            with open(metadata_file, 'r') as f:
                metadata = json.load(f)
            
            cache_time = datetime.fromisoformat(metadata.get('cache_time'))
            cache_age = datetime.now() - cache_time
            
            return {
                'cached': True,
                'cache_time': cache_time,
                'cache_age_hours': cache_age.total_seconds() / 3600,
                'rows': metadata.get('rows'),
                'columns': metadata.get('columns'),
                'valid': self.is_cache_valid(file_path)
            }
        except Exception:
            return None

class NumberLabel(QLabel):
    def __init__(self, number):
        super().__init__(str(number))
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setStyleSheet("""
            background-color: #f0f0f0;
            color: #333;
            border-radius: 12px;
            min-width: 24px;
            min-height: 24px;
            max-width: 24px;
            max-height: 24px;
            font-weight: bold;
            font-family: 'Segoe', sans-serif;
            font-size: 12px;
            box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
        """)

    def setActive(self, active):
        if active:
            self.setStyleSheet("""
                background-color: #008080;
                color: white;
                border-radius: 12px;
                min-width: 24px;
                min-height: 24px;
                max-width: 24px;
                max-height: 24px;
                font-weight: bold;
                font-family: 'Segoe', sans-serif;
                font-size: 12px;
                box-shadow: 0 2px 4px rgba(0, 0, 0, 0.2);
            """)
        else:
            self.setStyleSheet("""
                background-color: #f0f0f0;
                color: #333;
                border-radius: 12px;
                min-width: 24px;
                min-height: 24px;
                max-width: 24px;
                max-height: 24px;
                font-weight: bold;
                font-family: 'Segoe', sans-serif;
                font-size: 12px;
                box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
            """)

class GSMProcessorThread(QThread):
    progress_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str)
    bsc_names_signal = pyqtSignal(str)
    
    def __init__(self, bts_names, source_folder=None):
        super().__init__()
        self.source_folder = source_folder
        # Handle both string and list inputs for BTS names
        if isinstance(bts_names, list):
            self.bts_names = bts_names
        else:
            self.bts_names = [name.strip() for name in bts_names.split(',') if name.strip()]
        
    def run(self):
        try:
            self.progress_signal.emit("📄 Starting GSM template generation...")
            
            # Define column mapping (template_column: source_column)
            column_mapping = {
                "BSCName": "BSCName",
                "BTSName": "BTSName",
                "CellName": "CellName",
                "MCC": "MCC",
                "MNC": "MNC",
                "LAC": "LAC",
                "CI": "CI",
                "BCCHNo": "BCCHNO",
                "BCC": "BCC",
                "NCC": "NCC",
                "Latitude": "Latitude",
                "Longitude": "Longitude",
                "Azimuth": "Azimuth",
                "Outdoor": "IsOutdoor",
                "TCH": "TCH",
                "AntHeight": "AntennaHeight",
                "MechTilt": "MechanicalTilt",
                "ElecTilt": "ElectricalTilt"
            }
            
            # Validate source folder
            if not os.path.exists(self.source_folder):
                self.finished_signal.emit(False, f"Source folder not found: {self.source_folder}")
                return
                
            self.progress_signal.emit(f"🔍 Checking folder: {self.source_folder}")
            
            # Find all Excel files
            files = [f for f in os.listdir(self.source_folder) 
                    if f.startswith("GSM_GENEXCloud_Engineering_Parameter_ALL") 
                    and (f.endswith(".xlsx") or f.endswith(".xls"))]
            
            if not files:
                self.finished_signal.emit(False, "No matching GSM_GENEXCloud_Engineering_Parameter_ALL files found.")
                return
                
            latest_file = max(files, key=lambda x: os.path.getmtime(os.path.join(self.source_folder, x)))
            file_path = os.path.join(self.source_folder, latest_file)
            
            self.progress_signal.emit(f"📊 Processing GSM file: {latest_file}")
            
            # Read source workbook
            source_wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            source_ws = source_wb.active
            
            # Get header row and create column index mapping
            header_row = next(source_ws.rows)
            source_columns = {cell.value: idx for idx, cell in enumerate(header_row, 1)}
            
            # Validate required columns exist
            missing_columns = []
            for needed_col in column_mapping.values():
                if needed_col not in source_columns:
                    missing_columns.append(needed_col)
            
            if missing_columns:
                self.finished_signal.emit(False, f"Missing required columns: {', '.join(missing_columns)}")
                return
            
            # Get BTSNames from source - ensure they're uppercase for comparison
            original_bts_list = [name.strip().upper() for name in self.bts_names]
            
            bts_col_idx = source_columns['BTSName']
            bsc_col_idx = source_columns['BSCName']
            
            # Convert all rows to list of lists for easier processing
            all_rows = []
            bsc_names_dict = OrderedDict()
            found_bts_names = set()

            self.progress_signal.emit(f"🔍 Searching for {len(original_bts_list)} BTS names...")

            # First pass: Check for original BTS names
            for row in source_ws.iter_rows(min_row=2):
                row_data = [cell.value for cell in row]
                current_bts = row_data[bts_col_idx - 1]
                if current_bts:
                    current_bts_upper = str(current_bts).strip().upper()
                    if current_bts_upper in original_bts_list:
                        found_bts_names.add(current_bts_upper)
                        bsc_name = row_data[bsc_col_idx - 1]
                        if bsc_name:
                            bsc_name = str(bsc_name).strip()
                            if bsc_name not in bsc_names_dict:
                                bsc_names_dict[bsc_name] = True
                        all_rows.append(row_data)
            
            # Check which original BTS names weren't found
            not_found_bts = [bts for bts in original_bts_list if bts not in found_bts_names]
            
            # If any original BTS names weren't found, look for their C1 alternatives
            for bts_name in not_found_bts:
                if bts_name.endswith("M1"):
                    alternate_name = bts_name[:-2] + "C1"
                    self.progress_signal.emit(f"🔄 BTS {bts_name} not found. Looking for {alternate_name}")
                    
                    # Second pass: Check for alternate BTS names
                    source_wb_alt = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    source_ws_alt = source_wb_alt.active
                    
                    for row in source_ws_alt.iter_rows(min_row=2):
                        row_data = [cell.value for cell in row]
                        current_bts = row_data[bts_col_idx - 1]
                        if current_bts and str(current_bts).strip().upper() == alternate_name:
                            self.progress_signal.emit(f"✅ Found alternate: {alternate_name}")
                            bsc_name = row_data[bsc_col_idx - 1]
                            if bsc_name:
                                bsc_name = str(bsc_name).strip()
                                if bsc_name not in bsc_names_dict:
                                    bsc_names_dict[bsc_name] = True
                            all_rows.append(row_data)
                            break
                    source_wb_alt.close()
            
            source_wb.close()
            
            if bsc_names_dict:
                bsc_names_str = ", ".join(bsc_names_dict.keys())
                self.bsc_names_signal.emit(bsc_names_str)
                self.progress_signal.emit(f"🏢 Found BSCs: {bsc_names_str}")
            
            if not all_rows:
                self.finished_signal.emit(False, f"No records found for BTSNames: {', '.join(self.bts_names)}")
                return
            
            self.progress_signal.emit(f"📋 Found {len(all_rows)} matching GSM records")
            
            # Sort rows by CellName
            cellname_idx = source_columns['CellName'] - 1
            all_rows.sort(key=lambda x: str(x[cellname_idx]) if x[cellname_idx] is not None else '')
            
            # Prepare destination file
            dest_folder = os.path.join(os.path.expanduser('~'), 'Desktop', 'LTE_EP_Genex v2.1')
            dest_path = os.path.join(dest_folder, 'GSMEngineeringParameterTemplate.xlsx')
            
            os.makedirs(dest_folder, exist_ok=True)
            
            self.progress_signal.emit("📝 Creating GSM template file...")
            
            # Create or load destination workbook
            try:
                dest_wb = openpyxl.load_workbook(dest_path)
                dest_ws = dest_wb['GSM']
                # Clear existing data but keep headers
                for row in dest_ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.value = None
                self.progress_signal.emit("🔄 Cleared existing GSM template data")
            except FileNotFoundError:
                dest_wb = openpyxl.Workbook()
                dest_ws = dest_wb.active
                dest_ws.title = 'GSM'
                self.progress_signal.emit("🔄 Created new GSM template file")
            
            # Define cell style
            font = Font(name='Tahoma', size=11)
            alignment = Alignment(horizontal='center', vertical='center')
            no_border = Border(
                left=Side(style=None),
                right=Side(style=None),
                top=Side(style=None),
                bottom=Side(style=None)
            )
            
            # Define template column order
            template_columns = [
                "BSCName", "BTSName", "CellName", "MCC", "MNC", "LAC", "CI",
                "BCCHNo", "BCC", "NCC", "Latitude", "Longitude", "Azimuth",
                "Outdoor", "Sectorization", "TCH", "AntHeight", "MechTilt", "ElecTilt"
            ]
            
            # Write headers if new workbook or if headers are missing
            if dest_ws['A1'].value is None:
                self.progress_signal.emit("📝 Writing GSM template headers...")
                for col_idx, col_name in enumerate(template_columns, start=1):
                    header_cell = dest_ws.cell(row=1, column=col_idx)
                    header_cell.value = col_name
                    header_cell.font = Font(name='Tahoma', size=11, bold=True)
                    header_cell.alignment = alignment
            
            # Write sorted data
            self.progress_signal.emit(f"📊 Writing {len(all_rows)} GSM records...")
            for row_idx, row_data in enumerate(all_rows, start=2):
                for dest_col_idx, template_col in enumerate(template_columns, start=1):
                    cell = dest_ws.cell(row=row_idx, column=dest_col_idx)
                    
                    # Apply styling
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = no_border
                    
                    # Handle special columns
                    if template_col == "Sectorization":
                        cell.value = "SECTORIZE"
                    elif template_col == "Outdoor":
                        value = row_data[source_columns["IsOutdoor"] - 1]
                        cell.value = "YES" if value == 1 else "NO"
                    elif template_col == "TCH":
                        cell.value = None  # Leave TCH column blank
                    else:
                        # Map other columns
                        source_col = column_mapping.get(template_col)
                        if source_col and source_col in source_columns:
                            source_idx = source_columns[source_col] - 1
                            cell.value = row_data[source_idx]
            
            # Auto-adjust columns
            self.progress_signal.emit("🔧 Formatting GSM template...")
            for column in dest_ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        length = len(str(cell.value)) if cell.value else 0
                        if length > max_length:
                            max_length = length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                dest_ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            dest_wb.save(dest_path)
            dest_wb.close()
            self.progress_signal.emit(f"💾 GSM template saved: {dest_path}")
            
            bts_count = len(set(original_bts_list) & found_bts_names) + len([bts for bts in not_found_bts if bts.endswith("M1")])
            self.finished_signal.emit(True, f"GSM template generated successfully!\n{len(all_rows)} records processed for {bts_count} BTS sites.\nFile: GSMEngineeringParameterTemplate.xlsx")
                
        except Exception as e:
            self.finished_signal.emit(False, f"GSM processing error: {str(e)}")
            import traceback
            print(f"GSM Error details: {traceback.format_exc()}")

class ExcelFilterApp(QWidget):
    def __init__(self):
        super().__init__()
        self.start_time = QElapsedTimer()
        self.start_time.start()
        self.gsm_thread = None
        
        # Initialize database cache
        self.db_cache = DatabaseCache()
        self.available_bts_names = []
        self.available_bsc_names = []  # Add this to store BSC names
        self.initUI()
        self.center()

    def initUI(self):
        title = 'eNodeB & BTS Extractor (Unified Edition)'
        self.setWindowTitle(title)
        self.setWindowIcon(QIcon('myicon.ico'))

        # Set window properties
        font_metrics = QFontMetrics(self.font())
        title_width = font_metrics.horizontalAdvance(title)
        self.setGeometry(100, 100, title_width + 370, 110)
        self.setStyleSheet("""
            QWidget {
                background-color: #ffffff;
                color: #333333;
                font-family: 'Roboto', monospace;
            }
        """)

        # Main layout
        main_layout = QVBoxLayout()
        main_layout.setSpacing(4)
        main_layout.setContentsMargins(10, 10, 10, 10)
        self.setLayout(main_layout)

        # Element styles with Roboto
        button_style = """
            QPushButton {
                background-color: #008080;
                color: white;
                border: none;
                padding: 3px 6px;
                margin: 2px;
                border-radius: 1px;
                font-size: 12px;
                font-family: 'Roboto', monospace;
                font-weight: bold;
                min-height: 18px;
            }
            QPushButton:hover {
                background-color: #006666;
                font-weight: bold;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #ffffff;
                font-weight: bold;
            }
        """

        input_style = """
            QLineEdit {
                border: 1px solid #e0e0e0;
                border-radius: 3px;
                padding: 4px 8px;
                background: white;
                color: #333333;
                font-size: 11px;
                font-family: 'Roboto', monospace;
                min-height: 24px;
            }
            QLineEdit:focus {
                border: 1px solid #008080;
            }
        """

        label_style = """
            QLabel {
                color: #333333;
                font-family: 'Roboto', monospace;
                font-size: 11px;
                padding: 4px;
                background-color: #f7f7f7;
                border-radius: 3px;
            }
        """

        # Header label
        self.label = QLabel('Select LTE_GENEXCloud_Engineering_Parameter')
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.label.setStyleSheet(label_style)
        main_layout.addWidget(self.label)

        # Step 1: File Selection
        step1_layout = QHBoxLayout()
        step1_layout.setSpacing(4)
        self.step1_number = NumberLabel(1)
        step1_layout.addWidget(self.step1_number)
        self.btn = QPushButton('Select File')
        self.btn.setStyleSheet(button_style)
        self.btn.clicked.connect(self.selectFile)
        step1_layout.addWidget(self.btn)
        main_layout.addLayout(step1_layout)

        # Add filename display label
        self.filename_label = QLabel()
        self.filename_label.setStyleSheet("""
            QLabel {
                color: gray;
                font-family: 'Roboto', monospace;
                font-size: 10px;
                padding: 0px;
                margin: 0px;
                font-style: italic;
            }
        """)
        self.filename_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(self.filename_label)

        # Step 2: Filter Input
        step2_layout = QHBoxLayout()
        step2_layout.setSpacing(4)
        self.step2_number = NumberLabel(2)
        step2_layout.addWidget(self.step2_number)
        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText('Enter/paste eNodeB name(s) (auto-separated by spaces, tabs, newlines, or commas)')
        self.filter_input.setStyleSheet(input_style)
        self.filter_input.textChanged.connect(self.onFilterTextChanged)
        self.filter_input.returnPressed.connect(self.onEnterPressed)
        step2_layout.addWidget(self.filter_input)
        main_layout.addLayout(step2_layout)

        # Action Buttons - ONLY 5 buttons: Extract, Parse, Output, NEAccess, Reset
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(4)

        button_configs = [
            (3, 'Extract', self.applyFilter),
            (4, 'Parse', self.parsing),
            (5, 'Output', self.openOutputFolder),
            (6, 'NEAccess', self.neaccess),
            (7, 'Reset', self.reset)
        ]

        for step_num, btn_text, btn_func in button_configs:
            number_label = NumberLabel(step_num)
            buttons_layout.addWidget(number_label)
            setattr(self, f'step{step_num}_number', number_label)
            
            button = QPushButton(btn_text)
            button.setStyleSheet(button_style)
            button.setFixedWidth(85)
            button.clicked.connect(btn_func)
            buttons_layout.addWidget(button)
            if btn_text == 'Extract':
                self.filter_btn = button
            elif btn_text == 'Parse':
                self.parse_btn = button
            elif btn_text == 'Output':
                self.output_btn = button
            elif btn_text == 'NEAccess':
                self.neaccess_btn = button
            elif btn_text == 'Reset':
                self.reset_btn = button

        main_layout.addLayout(buttons_layout)

        # Progress Status Label
        self.progress_label = QLabel("")
        self.progress_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.progress_label.setStyleSheet("""
            QLabel {
                color: #008080;
                font-family: 'Roboto', monospace;
                font-size: 11px;
                font-weight: bold;
                padding: 4px 8px;
                background-color: #f0f8f8;
                border: 1px solid #d0e8e8;
                border-radius: 4px;
                margin: 2px 0px;
            }
        """)
        self.progress_label.hide()
        main_layout.addWidget(self.progress_label)

        # Status Section Layout
        status_layout = QHBoxLayout()
        status_layout.setSpacing(4)

        # Total count label
        self.total_count_label = QLabel("Count eNodeB: ")
        self.total_count_label.setStyleSheet("""
            QLabel {
                color: #666666;
                font-family: 'Roboto', monospace;
                font-size: 11px;
                padding: 1px;
                background-color: #f7f7f7;
                border-radius: 1px;
                margin-top: 2px;
            }
        """)
        self.total_count_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        status_layout.addWidget(self.total_count_label, 1)

        # Latlong label
        self.latlong_label = QLabel("Latitude/Longitude: ")
        self.latlong_label.setStyleSheet("""
            QLabel {
                color: #666666;
                font-family: 'Roboto', monospace;
                font-size: 11px;
                padding: 1px;
                background-color: #f7f7f7;
                border-radius: 1px;
                margin-top: 2px;
            }
        """)
        self.latlong_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        status_layout.addWidget(self.latlong_label, 1)

        # Height Check label
        self.height_check_label = QLabel("Antenna Height: ")
        self.height_check_label.setStyleSheet("""
            QLabel {
                color: #666666;
                font-family: 'Roboto', monospace;
                font-size: 11px;
                padding: 1px;
                background-color: #f7f7f7;
                border-radius: 1px;
                margin-top: 2px;
            }
        """)
        self.height_check_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        status_layout.addWidget(self.height_check_label, 1)

        # 2G Check label
        self.gsm_check_label = QLabel("2G: ")
        self.gsm_check_label.setStyleSheet("""
            QLabel {
                color: #666666;
                font-family: 'Roboto', monospace;
                font-size: 11px;
                padding: 1px;
                background-color: #f7f7f7;
                border-radius: 1px;
                margin-top: 2px;
            }
        """)
        self.gsm_check_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        status_layout.addWidget(self.gsm_check_label, 1)

        main_layout.addLayout(status_layout)    
            
        # Info Frame
        info_frame = QFrame()
        info_frame.setStyleSheet("""
            QFrame {
                background-color: #f7f7f7;
                border: none;
                border-radius: 3px;
            }
            QLabel {
                font-family: 'Roboto', monospace;
                font-size: 11px;
            }
        """)
        info_layout = QHBoxLayout(info_frame)
        info_layout.setSpacing(8)
        info_layout.setContentsMargins(12, 6, 12, 6)

        # eNodeB ID section
        left_container = QWidget()
        left_layout = QHBoxLayout(left_container)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(4)
        
        self.enodeb_id_label = QLabel("eNodeB ID:")
        self.enodeb_id_label.setStyleSheet("font-family: 'Roboto', monospace; color: #666666;")
        self.enodeb_id_value = QLabel("...")
        self.enodeb_id_value.setStyleSheet("font-family: 'Roboto', monospace; color: #008080;")
        
        left_layout.addWidget(self.enodeb_id_label)
        left_layout.addWidget(self.enodeb_id_value)
        left_layout.addStretch()

        # Vertical separator
        separator = QFrame()
        separator.setFrameShape(QFrame.Shape.VLine)
        separator.setStyleSheet("background-color: #e0e0e0;")

        # TAC section
        middle_container = QWidget()
        middle_layout = QHBoxLayout(middle_container)
        middle_layout.setContentsMargins(0, 0, 0, 0)
        middle_layout.setSpacing(4)
        
        self.tac_label = QLabel("TAC:")
        self.tac_label.setStyleSheet("font-family: 'Roboto', monospace; color: #666666;")
        self.tac_value = QLabel("...")
        self.tac_value.setStyleSheet("font-family: 'Roboto', monospace; color: #008080;")
        
        middle_layout.addWidget(self.tac_label)
        middle_layout.addWidget(self.tac_value)
        middle_layout.addStretch()

        # Second vertical separator
        separator2 = QFrame()
        separator2.setFrameShape(QFrame.Shape.VLine)
        separator2.setStyleSheet("background-color: #e0e0e0;")

        # Sub Region section
        right_container = QWidget()
        right_layout = QHBoxLayout(right_container)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(4)
        
        self.sub_region_label = QLabel("Sub Region:")
        self.sub_region_label.setStyleSheet("font-family: 'Roboto', monospace; color: #666666;")
        self.sub_region_value = QLabel("...")
        self.sub_region_value.setStyleSheet("font-family: 'Roboto', monospace; color: #008080;")
        
        right_layout.addWidget(self.sub_region_label)
        right_layout.addWidget(self.sub_region_value)
        right_layout.addStretch()

        # Add all sections to info layout
        info_layout.addWidget(left_container)
        info_layout.addWidget(separator)
        info_layout.addWidget(middle_container)
        info_layout.addWidget(separator2)
        info_layout.addWidget(right_container)

        main_layout.addWidget(info_frame)

        # Footer
        footer_layout = QHBoxLayout()
        footer_layout.setContentsMargins(2, 2, 2, 2)
        footer_layout.setSpacing(4)
        
        footer_style = """
            QLabel {
                color: #999999;
                font-size: 10px;
                padding: 0px;
                font-family: 'Roboto', monospace;
            }
        """
        self.uptime_label = QLabel()
        self.uptime_label.setStyleSheet(footer_style)
        footer_layout.addWidget(self.uptime_label)
        
        footer_text = QLabel("v6.4.4525 (Cache). With ❤️ by Fadzli Abdullah")
        footer_text.setStyleSheet(footer_style)
        footer_text.setAlignment(Qt.AlignmentFlag.AlignCenter)
        footer_layout.addWidget(footer_text)
        
        self.date_time_label = QLabel()
        self.date_time_label.setStyleSheet("""
            QLabel {
                color: #008080;
                font-size: 10px;
                padding: 0px;
                font-family: 'Roboto', monospace;
                font-weight: bold;
            }
        """)
        self.date_time_label.setAlignment(Qt.AlignmentFlag.AlignRight)
        footer_layout.addWidget(self.date_time_label)
        
        main_layout.addLayout(footer_layout)

        # Initialize properties
        self.df = None
        self.file_path = None
        self.output_folder = None
        self.filtered_df = None

        # Set up timer
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_date_time_and_uptime)
        self.timer.start(1000)
        self.update_date_time_and_uptime()

        # Apply Roboto font
        app_font = QFont("Roboto", 9)
        QApplication.setFont(app_font)

    def show_progress(self, message):
        """Show progress status message"""
        self.progress_label.setText(message)
        self.progress_label.show()
        QApplication.processEvents()

    def hide_progress(self):
        """Hide progress status message"""
        self.progress_label.hide()
        QApplication.processEvents()

    def update_date_time_and_uptime(self):
        # Update uptime
        elapsed_time = self.start_time.elapsed()
        hours, remainder = divmod(elapsed_time // 1000, 3600)
        minutes, seconds = divmod(remainder, 60)
        uptime_str = f"Uptime: {hours:02d}:{minutes:02d}:{seconds:02d}"
        self.uptime_label.setText(uptime_str)
        
        # Update CPU and RAM usage
        try:
            cpu_percent = psutil.cpu_percent(interval=0.1)
            memory = psutil.virtual_memory()
            ram_percent = memory.percent
            cpu_ram_str = f"CPU: {cpu_percent:.1f}% | RAM: {ram_percent:.1f}%"
            self.date_time_label.setText(cpu_ram_str)
        except Exception:
            self.date_time_label.setText("CPU: -- | RAM: --")

    def center(self):
        qr = self.frameGeometry()
        cp = QApplication.primaryScreen().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def selectFile(self):
        self.file_path, _ = QFileDialog.getOpenFileName(self, 'Fadzli Abdullah - Select File', '', "Excel files (*.xlsx *.xls)")
        if self.file_path:
            self.filename_label.setText(os.path.basename(self.file_path))
            
            # Check cache status
            cache_info = self.db_cache.get_cache_info(self.file_path)
            if cache_info and cache_info["valid"]:
                self.label.setText(f'Loading from cache ({cache_info["cache_age_hours"]:.1f}h old)...')
            else:
                self.label.setText('Loading file. Please wait.')
            QApplication.processEvents()
            QTimer.singleShot(100, self.loadFile)

    def loadFile(self):
        try:
            # Try to load from cache first
            if self.db_cache.is_cache_valid(self.file_path):
                self.show_progress("📦 Loading from cache...")
                self.df = self.db_cache.load_from_cache(self.file_path)
                
                if self.df is not None and 'eNodeB Name' in self.df.columns:
                    cache_info = self.db_cache.get_cache_info(self.file_path)
                    age_text = f"({cache_info['cache_age_hours']:.1f}h old)" if cache_info else ""
                    self.label.setText(f'File loaded from cache {age_text}. Enter eNodeB name and click Extract.')
                    self.hide_progress()
                    self.output_folder = os.path.dirname(self.file_path)
                    self.step1_number.setActive(False)
                    self.step2_number.setActive(True)
                    self.filter_input.setEnabled(True)
                    self.filter_input.clear()
                    for i in range(3, 8):
                        getattr(self, f'step{i}_number').setActive(False)
                    return
            
            # Load from Excel file if cache is not valid
            self.show_progress("📊 Loading Excel file...")
            self.df = pd.read_excel(self.file_path)
            if 'eNodeB Name' in self.df.columns:
                # Save to cache for future use
                self.show_progress("💾 Caching database for faster future loads...")
                self.db_cache.save_to_cache(self.file_path, self.df)
                
                self.label.setText('File loaded successfully (cached). Enter eNodeB name and click Extract.')
                self.hide_progress()
                self.output_folder = os.path.dirname(self.file_path)
                self.step1_number.setActive(False)
                self.step2_number.setActive(True)
                self.filter_input.setEnabled(True)
                self.filter_input.clear()
                for i in range(3, 8):
                    getattr(self, f'step{i}_number').setActive(False)
            else:
                self.hide_progress()
                self.label.setText('Fadzli Abdullah: No "eNodeB Name" column found.')
                self.df = None
        except Exception as e:
            self.hide_progress()
            self.label.setText(f'Fadzli Abdullah: {str(e)}')
            self.df = None

    def onFilterTextChanged(self, text):
        cursor_position = self.filter_input.cursorPosition()
        processed_names = self.process_enodeb_input(text)
        
        if len(processed_names) > 1:
            comma_separated = ', '.join(processed_names)
            if text != comma_separated:
                self.filter_input.textChanged.disconnect()
                self.filter_input.setText(comma_separated)
                new_cursor_position = min(cursor_position + (len(comma_separated) - len(text)), len(comma_separated))
                self.filter_input.setCursorPosition(new_cursor_position)
                self.filter_input.textChanged.connect(self.onFilterTextChanged)
        self.filter_input.returnPressed.connect(self.onEnterPressed)
        
        if len(processed_names) > 0 and all(len(name.strip()) >= 4 for name in processed_names):
            self.step3_number.setActive(True)
            self.filter_btn.setEnabled(True)
        else:
            self.step3_number.setActive(False)
            self.filter_btn.setEnabled(False)


    def onEnterPressed(self):
        """Handle Enter key press in filter input to trigger Extract"""
        # Only trigger if Extract button is enabled (valid input)
        if self.filter_btn.isEnabled():
            self.applyFilter()

    def process_enodeb_input(self, text):
        if not text.strip():
            return []
        
        import re
        processed_text = re.sub(r'[\n\t;]+', ',', text)
        processed_text = re.sub(r'\s+', ',', processed_text)
        processed_text = re.sub(r'\s*,\s*', ',', processed_text)
        processed_text = re.sub(r',+', ',', processed_text)
        processed_text = processed_text.strip(',')
        
        names = [name.strip() for name in processed_text.split(',') if name.strip()]
        
        # Truncate each eNodeB name to first 4 characters (4 LRD)
        names = [name[:4] for name in names]
        
        unique_names = []
        seen = set()
        for name in names:
            name_lower = name.lower()
            if name_lower not in seen:
                unique_names.append(name)
                seen.add(name_lower)
        
        return unique_names

    def applyFilter(self):
        if self.df is None:
            self.label.setText('Please load the LTE_GENEXCloud EP.')
            return
        
        filter_text = self.filter_input.text().strip()
        requested_enodebs = self.process_enodeb_input(filter_text)
        
        if not requested_enodebs or any(len(name.strip()) < 4 for name in requested_enodebs):
            self.label.setText('Each eNodeB name must be at least 4 chars long, 6 chars the best.')
            return
        
        self.show_progress("📄 Processing eNodeB extraction...")
        self.filter_btn.setEnabled(False)
        
        try:
            comma_separated_names = ', '.join(requested_enodebs)
            self.filter_input.setText(comma_separated_names)
            
            total_requested = len(requested_enodebs)
            
            self.show_progress("📂 Setting up output directory...")
            
            parent_dir = os.path.dirname(self.output_folder)
            existing_output_dir = None
            
            for root, dirs, files in os.walk(parent_dir):
                if "OUTPUT FILE" in dirs:
                    existing_output_dir = os.path.join(root, "OUTPUT FILE")
                    break
            
            if existing_output_dir:
                output_dir = existing_output_dir
            else:
                output_dir = os.path.join(self.output_folder, "OUTPUT FILE")
                if not os.path.exists(output_dir):
                    os.makedirs(output_dir)

            self.show_progress("🔍 Filtering eNodeB data...")
            
            enodeb_names = [name.strip().lower() for name in requested_enodebs]
            self.df['eNodeB Name'] = self.df['eNodeB Name'].astype(str)
            
            mask = self.df['eNodeB Name'].str.lower().apply(
                lambda x: any(
                    (name in x and 
                    (x.startswith(name) or x.endswith(name) or f"_{name}_" in f"_{x}_"))
                    for name in enodeb_names
                )
            )
            self.filtered_df = self.df[mask]

            # Initialize lists to store IDs and regions in order
            enodeb_ids = []
            tacs = []
            sub_regions = []
            
            for name in enodeb_names:
                matches = self.filtered_df[self.filtered_df['eNodeB Name'].str.lower().str.contains(name)]
                if not matches.empty:
                    first_match = matches.iloc[0]
                    enodeb_id = str(int(float(first_match['eNodeBID']))) if 'eNodeBID' in self.filtered_df.columns else 'N/A'
                    enodeb_ids.append(enodeb_id)
                    tac = str(int(float(first_match['TAC']))) if 'TAC' in self.filtered_df.columns else 'N/A'
                    tacs.append(tac)
                    sub_regions.append(str(first_match.iloc[45]) if len(first_match) > 45 else 'N/A')
                else:
                    enodeb_ids.append("N/A")
                    tacs.append("N/A")
                    sub_regions.append("N/A")

            if all(id == "N/A" for id in enodeb_ids):
                self.hide_progress()
                self.filter_btn.setEnabled(True)
                self.label.setText('The eNodeB name not found. Please check.')
                self.enodeb_id_value.setText("-")
                self.tac_value.setText("-")
                self.sub_region_value.setText("-")
                return

            combined_ids = '_'.join(enodeb_ids)
            combined_tacs = '_'.join(tacs)
            combined_regions = '*'.join(sub_regions)

            self.enodeb_id_value.setText(combined_ids)
            self.tac_value.setText(combined_tacs)
            self.sub_region_value.setText(combined_regions)

            self.show_progress("📊 Creating Excel file with formatting...")

            filename_text = '_'.join(requested_enodebs)
            save_path = os.path.join(output_dir, f"{filename_text.upper()}.xlsx")
            
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                self.filtered_df.to_excel(writer, index=False)
                worksheet = writer.sheets['Sheet1']
                
                mtnr_col = None
                for idx, col in enumerate(self.filtered_df.columns):
                    if col == 'MTNR Type':
                        mtnr_col = idx + 1
                        break
                
                color_32t32r = openpyxl.styles.PatternFill(
                    start_color='008B8B',
                    end_color='008B8B',
                    fill_type='solid'
                )
                color_8t8r = openpyxl.styles.PatternFill(
                    start_color='3B9C9C',
                    end_color='3B9C9C',
                    fill_type='solid'
                )
            
                if mtnr_col is not None:
                    for row in range(2, worksheet.max_row + 1):
                        cell_value = worksheet.cell(row=row, column=mtnr_col).value
                        if cell_value == '32T32R':
                            for col in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=row, column=col).fill = color_32t32r
                        elif cell_value == '8T8R':
                            for col in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=row, column=col).fill = color_8t8r

            extracted_count = len(self.filtered_df[self.filtered_df['eNodeB Name'].notna()]['eNodeB Name'].unique())
            self.total_count_label.setText(f"Count eNodeB: {extracted_count} of {total_requested}")

            self.show_progress("✅ Validating antenna height and coordinates...")

            if not self.filtered_df.empty:
                grouped = self.filtered_df.groupby('eNodeB Name')
                
                height_inconsistent = False
                latlong_inconsistent = False
                latlong_inconsistent_enodebs = []
                
                for name, group in grouped:
                    if group['AntennaHeight'].nunique() > 1:
                        height_inconsistent = True
                    
                    if group['Latitude'].nunique() > 1 or group['Longitude'].nunique() > 1:
                        latlong_inconsistent = True
                        latlong_inconsistent_enodebs.append(name)
                
                height_status = '<span style="background-color: #FF0000; color: #FFFFFF; padding: 2px 4px; border-radius: 2px;">NOT OKAY</span>' if height_inconsistent else '<span style="background-color: #008000; color: #FFFFFF; padding: 2px 4px; border-radius: 2px;">OKAY</span>'
                self.height_check_label.setText(f'Antenna Height: {height_status}')
                
                if latlong_inconsistent:
                    inconsistent_names = ', '.join(latlong_inconsistent_enodebs)
                    latlong_status = f'<span style="background-color: #FF0000; color: #FFFFFF; padding: 2px 4px; border-radius: 2px;">NOT OKAY</span> ({inconsistent_names})'
                else:
                    latlong_status = '<span style="background-color: #008000; color: #FFFFFF; padding: 2px 4px; border-radius: 2px;">OKAY</span>'
                
                self.latlong_label.setText(f'Latitude/Longitude: {latlong_status}')
            else:
                self.latlong_label.setText("Latitude/Longitude: -")

            self.output_folder = os.path.dirname(save_path)
            
            self.show_progress("📡 Checking 2G availability...")

            # Enhanced 2G availability check that also collects BTS names and BSC names
            gsm_status, bts_names, bsc_names = self.check_2g_availability_with_names_and_bsc(requested_enodebs)
            self.available_bts_names = bts_names
            self.available_bsc_names = bsc_names

            if gsm_status == "YES":
                status_color = "#008000"
                if bsc_names:
                    bsc_display = f" ({', '.join(bsc_names)})"
                    status_text = f'<span style="background-color: {status_color}; color: #FFFFFF; padding: 2px 4px; border-radius: 2px;">YES</span>{bsc_display}'
                else:
                    status_text = f'<span style="background-color: {status_color}; color: #FFFFFF; padding: 2px 4px; border-radius: 2px;">YES</span>'
            elif gsm_status == "NO":
                status_color = "#FF0000"
                status_text = f'<span style="background-color: {status_color}; color: #FFFFFF; padding: 2px 4px; border-radius: 2px;">NO</span>'
            elif "/" in gsm_status and "(" in gsm_status:
                parts = gsm_status.split(" (", 1)
                ratio_part = parts[0]
                available_part = f"({parts[1]}" if len(parts) > 1 else ""
                # Add BSC names if available
                if bsc_names:
                    bsc_display = f" [BSC: {', '.join(bsc_names)}]"
                    status_text = f'<span style="background-color: #666666; color: #FFFFFF; padding: 2px 4px; border-radius: 2px;">{ratio_part}</span> {available_part}{bsc_display}'
                else:
                    status_text = f'<span style="background-color: #666666; color: #FFFFFF; padding: 2px 4px; border-radius: 2px;">{ratio_part}</span> {available_part}'
            else:
                status_text = f'<span style="background-color: #666666; color: #FFFFFF; padding: 2px 4px; border-radius: 2px;">{gsm_status}</span>'

            self.gsm_check_label.setText(f'2G: {status_text}')
                        
            self.hide_progress()
            
            self.label.setText(f'Extracted eNodeB saved as: {os.path.basename(save_path)}')
            self.step3_number.setActive(False)
            self.step4_number.setActive(True)
            self.step5_number.setActive(True)
            self.step6_number.setActive(True)
            self.output_btn.setEnabled(True)
            self.parse_btn.setEnabled(True)
            self.neaccess_btn.setEnabled(True)
            self.step2_number.setActive(True)
            self.filter_btn.setEnabled(True)
            
        except Exception as e:
            self.hide_progress()
            self.filter_btn.setEnabled(True)
            
            self.label.setText(f'Error: {str(e)}')
            self.filtered_df = None
            self.enodeb_id_value.setText("-")
            self.sub_region_value.setText("-")
            self.total_count_label.setText("Count eNodeB: 0")
            self.height_check_label.setText("Antenna Height: -")
            self.latlong_label.setText("Latitude/Longitude: -")
            self.gsm_check_label.setText("2G: -")

    def check_2g_availability_with_names_and_bsc(self, enodeb_names):
        """
        Enhanced version that returns availability status, BTS names, and BSC names
        """
        try:
            if not self.file_path:
                return "File not loaded", [], []
            
            input_dir = os.path.dirname(self.file_path)
            gsm_files = [f for f in os.listdir(input_dir) if f.startswith('GSM_GENEXCloud_Engineering_Parameter_ALL_') and f.endswith('.xlsx')]
            
            if not gsm_files:
                return "GSM file not found", [], []
            
            gsm_file_path = os.path.join(input_dir, gsm_files[0])
            gsm_df = pd.read_excel(gsm_file_path)
            
            if 'BTSName' not in gsm_df.columns or 'BSCName' not in gsm_df.columns:
                return "Required columns not found", [], []
            
            # Convert to string and lowercase for comparison
            gsm_df['BTSName_lower'] = gsm_df['BTSName'].astype(str).str.lower()
            
            enodebs_with_bts = []
            enodebs_without_bts = []
            available_bts_names = []
            available_bsc_names = []
            
            for enodeb_name in enodeb_names:
                base_name = enodeb_name.strip().upper()
                
                if base_name.endswith('ML'):
                    site_code = base_name[:-2]
                elif base_name.endswith('CL'):
                    site_code = base_name[:-2]
                else:
                    site_code = base_name
                
                bts_variants = [
                    f"{site_code}M1",
                    f"{site_code}C1"
                ]
                
                found_variant = None
                found_bsc = None
                for variant in bts_variants:
                    matching_rows = gsm_df[gsm_df['BTSName_lower'] == variant.lower()]
                    if not matching_rows.empty:
                        found_variant = variant
                        # Get the BSC name from the first matching row
                        found_bsc = str(matching_rows.iloc[0]['BSCName']).strip()
                        break
                
                if found_variant:
                    enodebs_with_bts.append(enodeb_name)
                    available_bts_names.append(found_variant)
                    if found_bsc and found_bsc not in available_bsc_names:
                        available_bsc_names.append(found_bsc)
                else:
                    enodebs_without_bts.append(enodeb_name)
            
            if len(enodebs_without_bts) == 0:
                return "YES", available_bts_names, available_bsc_names
            elif len(enodebs_with_bts) == 0:
                return "NO", [], []
            else:
                available_names = ', '.join(enodebs_with_bts)
                return f"{len(enodebs_with_bts)}/{len(enodeb_names)} ({available_names})", available_bts_names, available_bsc_names
                
        except Exception as e:
            return f"Error: {str(e)[:20]}...", [], []

    def check_2g_availability_with_names(self, enodeb_names):
        """
        Enhanced version that returns both availability status and BTS names
        """
        try:
            if not self.file_path:
                return "File not loaded", []
            
            input_dir = os.path.dirname(self.file_path)
            gsm_files = [f for f in os.listdir(input_dir) if f.startswith('GSM_GENEXCloud_Engineering_Parameter_ALL_') and f.endswith('.xlsx')]
            
            if not gsm_files:
                return "GSM file not found", []
            
            gsm_file_path = os.path.join(input_dir, gsm_files[0])
            gsm_df = pd.read_excel(gsm_file_path)
            
            if 'BTSName' not in gsm_df.columns:
                return "BTSName column not found", []
            
            bts_names = gsm_df['BTSName'].astype(str).str.lower().tolist()
            
            enodebs_with_bts = []
            enodebs_without_bts = []
            available_bts_names = []
            
            for enodeb_name in enodeb_names:
                base_name = enodeb_name.strip().upper()
                
                if base_name.endswith('ML'):
                    site_code = base_name[:-2]
                elif base_name.endswith('CL'):
                    site_code = base_name[:-2]
                else:
                    site_code = base_name
                
                bts_variants = [
                    f"{site_code}M1",
                    f"{site_code}C1"
                ]
                
                found_variant = None
                for variant in bts_variants:
                    if variant.lower() in bts_names:
                        found_variant = variant
                        break
                
                if found_variant:
                    enodebs_with_bts.append(enodeb_name)
                    available_bts_names.append(found_variant)
                else:
                    enodebs_without_bts.append(enodeb_name)
            
            if len(enodebs_without_bts) == 0:
                return "YES", available_bts_names
            elif len(enodebs_with_bts) == 0:
                return "NO", []
            else:
                available_names = ', '.join(enodebs_with_bts)
                return f"{len(enodebs_with_bts)}/{len(enodeb_names)} ({available_names})", available_bts_names
                
        except Exception as e:
            return f"Error: {str(e)[:20]}...", []

    def parsing(self):
        # Set default directory to "Output File" folder if it exists
        default_dir = os.path.join(self.output_folder, "OUTPUT FILE") if hasattr(self, 'output_folder') and self.output_folder else ''
        
        if os.path.exists(default_dir):
            input_file_path, _ = QFileDialog.getOpenFileName(self, 'Fadzli Abdullah - Select File', default_dir, "Excel files (*.xlsx *.xls)")
        else:
            input_file_path, _ = QFileDialog.getOpenFileName(self, 'Fadzli Abdullah - Select File', '', "Excel files (*.xlsx *.xls)")
        
        if not input_file_path:
            return

        output_file_path = r"C:\Users\mWX1318105\Desktop\LTE_EP_Genex v2.1\LTEEngineeringParameterTemplate.xlsx"

        try:
            if not os.path.exists(output_file_path):
                self.show_styled_message_box("Error", "Template file not found. Please check the path.", QMessageBox.Icon.Critical)
                return

            try:
                test_wb = openpyxl.load_workbook(output_file_path)
                test_wb.close()
            except PermissionError:
                self.show_styled_message_box("Error", "Template file is open. Please close it and try again.", QMessageBox.Icon.Critical)
                return

            # Start LTE parsing
            self.show_progress("📊 Processing LTE template...")
            
            input_workbook = openpyxl.load_workbook(input_file_path)
            input_sheet = input_workbook.active

            output_workbook = openpyxl.load_workbook(output_file_path)
            output_sheet = output_workbook["LTE"]

            try:
                # Store the background colors of rows
                row_colors = {}
                for row in range(2, input_sheet.max_row + 1):
                    first_cell = input_sheet.cell(row=row, column=1)
                    if first_cell.fill and first_cell.fill.start_color.rgb != '00000000':
                        row_colors[row] = first_cell.fill.start_color.rgb

                column_mappings = {
                    "eNodeB Name": "eNodeB Name",
                    "CellName": "Cell Name",
                    "eNodeBID": "eNodeB ID",
                    "LocalCellID": "Local Cell ID",
                    "CellID": "Cell ID",
                    "PCI": "PCI",
                    "TAC": "TAC",
                    "Latitude": "Latitude",
                    "Longitude": "Longitude",
                    "DlEarfcn": "DLEARFCN",
                    "AntennaHeight": "Height",
                    "MechanicalTilt": "Mechanical Downtilt",
                    "ElectricalTilt": "Electrical Downtilt",
                    "Pattern": "Antenna Pattern",
                    "BeamWidth": "Beamwidth",
                    "IsOutDoor": "isOutdoor",
                    "On Air / Planned / Decomm / Dummy / Temporary / MBTS": "OnAir",
                    "MCC": "MCC",
                    "MNC": "MNC",
                    "Duplex Model": "Duplex Model",
                    "FrequencyBandS": "FreqBand",
                    "Cell active state": "Active",
                    "Operator": "Operator",
                    "MME": "Vendor",
                    "MTNR Type": "MTNR Type"
                }

                # Clear existing data in output sheet
                for row in output_sheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.value = None
                        cell.fill = openpyxl.styles.PatternFill(fill_type=None)
                
                # Column handling logic (same as original)
                mtnr_col = None
                azimuth_col = None
                azimuth_logical_col = None
                for idx, col in enumerate(input_sheet[1], 1):
                    if col.value == 'MTNR Type':
                        mtnr_col = idx
                    elif col.value == 'Azimuth':
                        azimuth_col = idx
                    elif col.value == 'Azimuth (Logical)':
                        azimuth_logical_col = idx

                output_azimuth_col = None
                for idx, col in enumerate(output_sheet[1], 1):
                    if col.value == 'Azimuth':
                        output_azimuth_col = idx
                        break

                last_row = input_sheet.max_row
                output_row_mapping = {}

                # Handle regular column mappings
                for col in input_sheet[1]:
                    input_column_name = col.value
                    if input_column_name in column_mappings and input_column_name != 'Azimuth (Logical)':
                        output_column_name = column_mappings[input_column_name]
                        input_column = [cell.value for cell in input_sheet[col.column_letter][1:last_row]]
                        output_column = next((cell for cell in output_sheet[1] if cell.value == output_column_name), None)
                        
                        if output_column:
                            for i, value in enumerate(input_column, start=2):
                                output_cell = output_sheet.cell(row=i, column=output_column.column)
                                output_cell.value = value
                                output_row_mapping[i] = i
                        
                # Special handling for Azimuth based on MTNR Type
                if mtnr_col and azimuth_col and azimuth_logical_col and output_azimuth_col:
                    for row in range(2, last_row + 1):
                        mtnr_value = input_sheet.cell(row=row, column=mtnr_col).value
                        if mtnr_value == '32T32R':
                            azimuth_value = input_sheet.cell(row=row, column=azimuth_col).value
                        else:
                            azimuth_value = input_sheet.cell(row=row, column=azimuth_logical_col).value
                        
                        output_sheet.cell(row=row, column=output_azimuth_col).value = azimuth_value

                # Apply highlighting to output rows
                for input_row, color_rgb in row_colors.items():
                    if input_row in output_row_mapping:
                        output_row = output_row_mapping[input_row]
                        new_fill = openpyxl.styles.PatternFill(
                            start_color=color_rgb[2:] if color_rgb.startswith('0x') else color_rgb,
                            end_color=color_rgb[2:] if color_rgb.startswith('0x') else color_rgb,
                            fill_type='solid'
                        )
                        for col in range(1, output_sheet.max_column + 1):
                            output_sheet.cell(row=output_row, column=col).fill = new_fill

                # Process special columns and values
                for row in range(2, output_sheet.max_row + 1):
                    if output_sheet.cell(row=row, column=1).value:
                        # Site Code processing
                        cell_value = output_sheet.cell(row=row, column=1).value
                        if isinstance(cell_value, str):
                            output_sheet.cell(row=row, column=2).value = cell_value[:4]

                        # Copy Cell ID to another column
                        output_sheet.cell(row=row, column=5).value = output_sheet.cell(row=row, column=6).value

                        # IsOutDoor conversion
                        cell = output_sheet.cell(row=row, column=19)
                        cell.value = "YES" if cell.value == 1 else "NO" if cell.value == 0 else cell.value

                        # Set Sectorize
                        output_sheet.cell(row=row, column=24).value = "SECTORIZE"

                        # Copy vendor data
                        output_sheet.cell(row=row, column=17).value = output_sheet.cell(row=row, column=50).value

                        # Set Macro/Micro based on IsOutDoor
                        cell = output_sheet.cell(row=row, column=23)
                        value = output_sheet.cell(row=row, column=19).value
                        cell.value = "Macro" if value == "YES" else "Micro" if value == "NO" else cell.value

                        # OnAir status conversion
                        cell = output_sheet.cell(row=row, column=22)
                        cell.value = "NO" if cell.value == "Plan" else "YES" if cell.value == "On Air" else cell.value

                        # Cell active state conversion
                        cell = output_sheet.cell(row=row, column=48)
                        if cell.value in ["CELL_ACTIVE", "CELL_DEACTIVE"]:
                            cell.value = "YES"

                # Apply standard formatting
                font = Font(name="Tahoma", size=11)
                alignment = Alignment(vertical="center", horizontal="center")
                no_border = Border(
                    left=Side(style=None), 
                    right=Side(style=None), 
                    top=Side(style=None), 
                    bottom=Side(style=None)
                )
                
                for row in output_sheet.iter_rows(min_row=1, max_row=output_sheet.max_row, min_col=1, max_col=55):
                    for cell in row:
                        cell.font = font
                        cell.alignment = alignment
                        cell.border = no_border

                # Save and close workbooks
                output_workbook.save(output_file_path)
                output_workbook.close()
                input_workbook.close()

                self.show_progress("✅ LTE template completed!")

                # Check if GSM template should be auto-generated
                if self.available_bts_names and self.file_path:
                    self.show_progress("📄 Auto-generating GSM template...")
                    
                    # Start GSM processing thread automatically
                    bts_names_string = ','.join(self.available_bts_names)
                    source_folder = os.path.dirname(self.file_path)
                    self.gsm_thread = GSMProcessorThread(bts_names_string, source_folder)
                    self.gsm_thread.progress_signal.connect(self.show_progress)
                    self.gsm_thread.finished_signal.connect(self.auto_gsm_completed)
                    self.gsm_thread.bsc_names_signal.connect(self.update_bsc_display)
                    self.gsm_thread.start()
                else:
                    # No GSM data available, just show LTE completion
                    self.parsing_completed_lte_only(output_file_path)

            finally:
                try:
                    output_workbook.close()
                    input_workbook.close()
                except:
                    pass

        except Exception as e:
            self.hide_progress()
            error_msg = f"An error occurred during parsing:\n\n{str(e)}\n\nTraceback:\n{traceback.format_exc()}"
            self.show_styled_message_box("Error", error_msg, QMessageBox.Icon.Critical)

    def auto_gsm_completed(self, success, message):
        """Handle automatic GSM generation completion during parsing"""
        self.hide_progress()
        
        if success:
            # Both LTE and GSM templates generated successfully
            winsound.Beep(750, 300)
            result = self.show_styled_message_box(
                "Templates Generated", 
                f"Both LTE and GSM templates generated successfully!\n\nLTE: LTEEngineeringParameterTemplate.xlsx\nGSM: GSMEngineeringParameterTemplate.xlsx\n\nClick OK to open the LTE template file."
            )
            
            if result is not None and result == QMessageBox.StandardButton.Ok:
                def open_lte_file():
                    try:
                        lte_output_file_path = r"C:\Users\mWX1318105\Desktop\LTE_EP_Genex v2.1\LTEEngineeringParameterTemplate.xlsx"
                        if sys.platform == 'win32':
                            os.startfile(lte_output_file_path)
                        elif sys.platform == 'darwin':
                            subprocess.run(['open', lte_output_file_path], check=True)
                        else:
                            subprocess.run(['xdg-open', lte_output_file_path], check=True)
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Could not open the LTE template file: {str(e)}")

                QTimer.singleShot(100, open_lte_file)
        else:
            # GSM generation failed, but LTE was successful
            result = self.show_styled_message_box(
                "Partial Success", 
                f"LTE template generated successfully, but GSM generation failed:\n{message}\n\nClick OK to open the LTE template file."
            )
            
            if result is not None and result == QMessageBox.StandardButton.Ok:
                def open_lte_file():
                    try:
                        lte_output_file_path = r"C:\Users\mWX1318105\Desktop\LTE_EP_Genex v2.1\LTEEngineeringParameterTemplate.xlsx"
                        if sys.platform == 'win32':
                            os.startfile(lte_output_file_path)
                        elif sys.platform == 'darwin':
                            subprocess.run(['open', lte_output_file_path], check=True)
                        else:
                            subprocess.run(['xdg-open', lte_output_file_path], check=True)
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Could not open the LTE template file: {str(e)}")

                QTimer.singleShot(100, open_lte_file)

    def parsing_completed_lte_only(self, output_file_path):
        """Handle completion when only LTE template is generated"""
        self.hide_progress()
        winsound.Beep(750, 300)
        result = self.show_styled_message_box("LTE Template Generated", "LTE template parsing completed.\n\nClick OK to open the template file.")
        
        if result is not None and result == QMessageBox.StandardButton.Ok:
            def open_file():
                try:
                    if sys.platform == 'win32':
                        os.startfile(output_file_path)
                    elif sys.platform == 'darwin':
                        subprocess.run(['open', output_file_path], check=True)
                    else:
                        subprocess.run(['xdg-open', output_file_path], check=True)
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Could not open the Output File: {str(e)}")

            QTimer.singleShot(100, open_file)

    def update_bsc_display(self, bsc_names):
        """Update the progress label with BSC names during GSM processing"""
        self.show_progress(f"🏢 BSC Names: {bsc_names}")
                
    def show_styled_message_box(self, title, message, icon_type=QMessageBox.Icon.Information):
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle(title)
        msg_box.setText(message)
        
        if icon_type == QMessageBox.Icon.Information:
            svg_string = """
            <svg width="40" height="40" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
                <circle cx="12" cy="12" r="10" fill="#008080"/>
                <circle cx="12" cy="8" r="1.5" fill="white"/>
                <rect x="11" y="11" width="2" height="7" fill="white"/>
            </svg>
            """
            pixmap = QPixmap()
            pixmap.loadFromData(svg_string.encode('utf-8'), 'SVG')
            msg_box.setIconPixmap(pixmap)
        else:
            msg_box.setIcon(icon_type)
        
        msg_box.setStyleSheet("""
            QMessageBox {
                background-color: #FFFFFF;
                font-family: 'Segoe', sans-serif;
            }
            QMessageBox QLabel {
                color: #333333;
                font-size: 12px;
            }
            QPushButton {
                background-color: #008080;
                color: white;
                border: none;
                padding: 6px 12px;
                margin: 4px;
                border-radius: 4px;
                font-size: 12px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #006666;
            }
        """)
        
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        result = msg_box.exec()
        
        if result == 0:
            return None
        return result

    def openOutputFolder(self):
        try:
            if not self.output_folder:
                self.label.setText('Please select and extract file first.')
                return

            if os.path.exists(self.output_folder):
                if "OUTPUT FILE" in os.path.basename(self.output_folder):
                    output_dir = self.output_folder
                else:
                    output_dir = os.path.join(self.output_folder, "OUTPUT FILE")
            else:
                parent_dir = os.path.dirname(self.output_folder)
                output_dir = None
                for root, dirs, _ in os.walk(parent_dir):
                    if "OUTPUT FILE" in dirs:
                        output_dir = os.path.join(root, "OUTPUT FILE")
                        break
            
            if output_dir and os.path.exists(output_dir):
                if sys.platform == 'win32':
                    os.startfile(output_dir)
                elif sys.platform == 'darwin':
                    subprocess.run(['open', output_dir])
                else:
                    subprocess.run(['xdg-open', output_dir])
            else:
                self.label.setText('OUTPUT FILE folder does not exist. Please extract file first.')
        except Exception as e:
            self.label.setText(f'Error opening output folder: {str(e)}')

    def update_template_file(self, template_path, enodeb_names):
        """Helper function to update a single template file with eNodeB names"""
        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook['Sheet1']

        # Clear existing data in column A, starting from row 2
        for row in sheet['A2:A' + str(sheet.max_row)]:
            for cell in row:
                cell.value = None

        # Set default font style for the entire sheet
        default_font = Font(name="Tahoma", size=11)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.font = default_font

        # Write eNodeB names with all variations to column A, starting from row 2
        row_index = 2
        for name in enodeb_names:
            variations = [
                name,
                f"{name}_INT",
                f"{name}ML",
                f"{name}ML_INT",
                f"{name}CL",
                f"{name}CL_INT",
                f"{name}BL",
                f"{name}BL_INT"
            ]
            
            for variation in variations:
                cell = sheet.cell(row=row_index, column=1, value=variation)
                cell.font = default_font
                row_index += 1

        workbook.save(template_path)
        workbook.close()

    def neaccess(self):
        try:
            if not hasattr(self, 'filtered_df') or self.filtered_df is None or self.filtered_df.empty:
                self.show_styled_message_box("Reminder", "You must extract eNodeB Names first.")
                self.label.setText('Please extract the eNodeB Name.')
                return

            # Find the OUTPUT FILE folder
            output_dir = None
            if self.output_folder:
                if "OUTPUT FILE" in os.path.basename(self.output_folder):
                    output_dir = self.output_folder
                else:
                    immediate_output_dir = os.path.join(self.output_folder, "OUTPUT FILE")
                    if os.path.exists(immediate_output_dir):
                        output_dir = immediate_output_dir
                    else:
                        parent_dir = os.path.dirname(self.output_folder)
                        for root, dirs, _ in os.walk(parent_dir):
                            if "OUTPUT FILE" in dirs:
                                output_dir = os.path.join(root, "OUTPUT FILE")
                                break

            if not output_dir or not os.path.exists(output_dir):
                self.show_styled_message_box("Folder Not Found", "OUTPUT FILE folder does not exist. Please extract file first.")
                self.label.setText('OUTPUT FILE folder does not exist. Please extract file first.')
                return

            excel_files = [f for f in os.listdir(output_dir) if f.endswith('.xlsx')]
            
            if excel_files:
                latest_file = max([os.path.join(output_dir, f) for f in excel_files], 
                                key=os.path.getctime)
                filename_without_ext = os.path.splitext(os.path.basename(latest_file))[0]
                enodeb_names = [name.strip() for name in filename_without_ext.replace('_', ',').split(',') if name.strip()]
                enodeb_names = list(dict.fromkeys(enodeb_names))
            else:
                self.show_styled_message_box("File Not Found", "No extracted file found in OUTPUT FILE folder.")
                self.label.setText('No extracted file found in OUTPUT FILE folder.')
                return
                    
            base_dir = os.path.dirname(sys.argv[0])
            
            template_files = [
                ("NE(Device)_Batch_Select_Template.xlsx", "NE(Device)_Batch_Select_Template"),
                ("DataSubscribe_Batch_Select_Template.xlsx", "DataSubscribe_Batch_Select_Template")
            ]
            
            updated_templates = []
            primary_template_path = None
            
            for filename, display_name in template_files:
                template_path = os.path.join(base_dir, filename)
                
                if not os.path.exists(template_path):
                    self.show_styled_message_box("File Not Found", f"{filename} not found in the script directory.")
                    self.label.setText(f'{filename} not found in the script directory.')
                    continue
                
                try:
                    self.update_template_file(template_path, enodeb_names)
                    updated_templates.append(display_name)
                    
                    if primary_template_path is None:
                        primary_template_path = template_path
                        
                except Exception as e:
                    self.show_styled_message_box("Error", f"An error occurred while updating {filename}: {str(e)}", QMessageBox.Icon.Critical)
                    continue

            filename_text = "_".join(enodeb_names)

            if updated_templates:
                if len(updated_templates) == 2:
                    self.label.setText(f'Templates updated: NE(Device) & DataSubscribe - {filename_text}')
                else:
                    self.label.setText(f'Template updated: {updated_templates[0]} - {filename_text}')
                
                self.step6_number.setActive(False)

                if primary_template_path:
                    if sys.platform == 'win32':
                        os.startfile(primary_template_path)
                    elif sys.platform == 'darwin':
                        subprocess.run(['open', primary_template_path])
                    else:
                        subprocess.run(['xdg-open', primary_template_path])
            else:
                self.show_styled_message_box("Error", "No template files were successfully updated.", QMessageBox.Icon.Critical)
                self.label.setText('No template files were successfully updated.')

        except Exception as e:
            self.show_styled_message_box("Error", f"An error occurred: {str(e)}", QMessageBox.Icon.Critical)
            self.label.setText('An error occurred while updating template files.')

    def reset(self):
        try:
            # Reset all display values
            self.enodeb_id_value.setText("...")
            self.tac_value.setText("...")
            self.sub_region_value.setText("...")
            self.total_count_label.setText("Count eNodeB: ")
            self.height_check_label.setText("Antenna Height: ")
            self.latlong_label.setText("Latitude/Longitude: ")
            self.gsm_check_label.setText("2G: ")
            self.label.setText("Select LTE_GENEXCloud_Engineering_Parameter")
            self.filter_input.clear()
            
            # Clear available BTS names and BSC names
            self.available_bts_names = []
            self.available_bsc_names = []
            
            self.hide_progress()
            
            # Reset step indicators (except step 1 since we're keeping the loaded file)
            for i in range(2, 8):
                number_label = getattr(self, f'step{i}_number')
                number_label.setActive(False)
            
            self.step1_number.setActive(False)
            self.step2_number.setActive(True)
            
            # Reset button states
            self.filter_btn.setEnabled(False)
            self.parse_btn.setEnabled(False)
            self.output_btn.setEnabled(False)
            self.neaccess_btn.setEnabled(False)
            self.reset_btn.setEnabled(True)
            self.filter_input.setEnabled(True)
            
            # Clear only the filtered dataframe
            self.filtered_df = None
            
            # Restore original window size
            title = 'eNodeB & BTS Extractor (Auto-Gen Edition)'
            font_metrics = QFontMetrics(self.font())
            title_width = font_metrics.horizontalAdvance(title)
            original_width = title_width + 370
            original_height = 110
            
            self.setFixedSize(original_width, original_height)
            self.resize(original_width, original_height)
            self.setMinimumSize(0, 0)
            self.setMaximumSize(16777215, 16777215)
            self.center()
            
            # Find and delete extracted files in Output Folder
            output_dir = None
            if self.output_folder:
                if "OUTPUT FILE" in os.path.basename(self.output_folder):
                    output_dir = self.output_folder
                else:
                    immediate_output_dir = os.path.join(self.output_folder, "OUTPUT FILE")
                    if os.path.exists(immediate_output_dir):
                        output_dir = immediate_output_dir
                    else:
                        parent_dir = os.path.dirname(self.output_folder)
                        for root, dirs, _ in os.walk(parent_dir):
                            if "OUTPUT FILE" in dirs:
                                output_dir = os.path.join(root, "OUTPUT FILE")
                                break
            
            files_deleted = False
            if output_dir and os.path.exists(output_dir):
                for file in os.listdir(output_dir):
                    if file.endswith('.xlsx'):
                        file_path = os.path.join(output_dir, file)
                        try:
                            os.remove(file_path)
                            files_deleted = True
                        except Exception as e:
                            self.show_styled_message_box(
                                "Warning",
                                f"Could not delete file {file}: {str(e)}",
                                QMessageBox.Icon.Warning
                            )
            
            if files_deleted:
                self.show_styled_message_box(
                    "Fadzli Abdullah",
                    "Reset complete. Extracted file(s) successfully deleted."
                )
            else:
                self.show_styled_message_box(
                    "Fadzli Abdullah",
                    "Reset complete. No files to delete."
                )
            
        except Exception as e:
            self.show_styled_message_box(
                "Error",
                f"An error occurred during reset: {str(e)}",
                QMessageBox.Icon.Critical
            )

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ExcelFilterApp()
    ex.show()
    sys.exit(app.exec())