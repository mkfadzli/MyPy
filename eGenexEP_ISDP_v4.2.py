import sys
import os
import winsound
from PyQt6.QtWidgets import (QApplication, QMainWindow, QPushButton, QVBoxLayout,
                            QWidget, QFileDialog, QMessageBox, QComboBox, QLabel,
                            QHBoxLayout, QStatusBar, QFrame, QLineEdit, QProgressBar)
from PyQt6.QtCore import Qt, QSettings, QTimer, QDateTime, QPoint
from PyQt6.QtGui import QFont, QIcon, QColor, QPalette
import pandas as pd
from fuzzywuzzy import fuzz
import openpyxl
from openpyxl.styles import Font, Alignment

def resource_path(relative_path):
    """ Get absolute path to resource """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base_path, relative_path)

class ModernExcelProcessor(QMainWindow):
    def __init__(self):
        super().__init__()
        # Set application icon
        icon_path = resource_path('myicon.ico')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        self.filtered_df = None
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        
        # Initialize settings
        self.settings = QSettings('ModernExcelProcessor', 'App')
        self.output_folder = self.settings.value('output_folder', os.path.expanduser("C:\\Users\\mWX1318105\\Desktop\\LTE_EP_Genex v2.1\\Engineering Parameter\\OUTPUT FILE"))
        self.selected_file = None
        
        # Initialize data mappings
        self.column_mapping = {
            'eNodeB Name': ('eNodeB Name', None),
            'Site Name': ('4LRD', None),
            'Cell Name': ('CellName', None),
            'eNodeB ID': ('eNodeBID', None),
            'SectorEqmId': ('SectorId', None),
            'Local Cell ID': ('LocalCellID', None),
            'Cell ID': ('CellId', None),
            'PCI': ('PCI', None),
            'TAC': ('TAC', None),
            'Latitude': ('Latitude', None),
            'Longitude': ('Longitude', None),
            'Height': ('AntHeight(m)', None),
            'Azimuth': ('Azimuth', None),
            'Mechanical Downtilt': ('Plan Mechanical Tilt', None),
            'Electrical Downtilt': ('Plan RET (E-Tilt)', None),
            'Antenna Pattern': ('Antenna Type', None),
            'DLEARFCN': ('DlEarfcn', None),
            'MCC': ('MCC', None),
            'MNC': ('MNC', None),
            'Duplex Model': ('FddTddInd', None),
            'FreqBand': ('FrequencyBand', None),
            'MTNR Type': ('TxRxMode', None),
            'Sub Region': ('KV BORDER', None)  # Add mapping for Sub Region to use KV
        }
        
        self.static_values = {
            'Antenna Vendor': ('Huawei', None),
            'Beamwidth': ('65', None),
            'isOutdoor': ('YES', None),
            'ULEARFCN': ('', None),
            'OnAir': ('YES', None),
            'Site Type': ('Macro', None),
            'Sectorization': ('SECTORIZE', None),
            'Active': ('YES', None),
            'Operator': ('MAXIS', None),
            'Vendor': ('Huawei', None),
            'High-speed Rail Way Side Cell': ('', None),
            'High-speed Rail Station': ('', None),
            'High-speed Rail Way Cell': ('', None),
            'High-speed Rail Way Exclusive Cell': ('', None)
        }
        
        self.init_ui()
        
        # Setup uptime tracking
        self.start_time = QDateTime.currentDateTime()
        self.uptime_timer = QTimer()
        self.uptime_timer.timeout.connect(self.update_uptime)
        self.uptime_timer.start(1000)
    
    def init_ui(self):
        # Main container
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        
        # Create main layout with styling
        layout = QVBoxLayout(main_widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Create a frame for the entire window
        container = QFrame()
        container.setObjectName("mainContainer")
        container.setStyleSheet("""
            #mainContainer {
                background-color: white;
                border-radius: 10px;
                border: 1px solid #e0e0e0;
            }
        """)
        
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(0)
        
        # Add title bar
        title_bar = self.create_title_bar()
        container_layout.addWidget(title_bar)
        
        # Content area
        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(4, 4, 4, 4)
        content_layout.setSpacing(4)
        
        # File selection area
        file_frame = self.create_file_section()
        content_layout.addWidget(file_frame)
        
        # Filter section
        filter_frame = self.create_filter_section()
        content_layout.addWidget(filter_frame)
        
        # Status display section
        status_frame = self.create_status_section()
        content_layout.addWidget(status_frame)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedHeight(4)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setVisible(False)
        content_layout.addWidget(self.progress_bar)
        
        # Footer
        footer = self.create_footer()
        content_layout.addWidget(footer)
        
        container_layout.addWidget(content)
        layout.addWidget(container)
        
        # Update window dimensions for the new layout
        self.setFixedSize(600, 260)
        self.center_window()
        
        # Apply global stylesheet
        self.apply_global_stylesheet()

    def create_title_bar(self):
        title_bar = QWidget()
        title_bar.setFixedHeight(40)
        title_bar.setStyleSheet("""
            QWidget {
                background: transparent;
            }
        """)
        
        # Create inner container for the teal background with proper sizing
        inner_container = QWidget(title_bar)
        inner_container.setStyleSheet("""
            QWidget {
                background: #008080;
                border-top-left-radius: 10px;
                border-top-right-radius: 10px;
            }
        """)
        # Important: Update geometry in resizeEvent
        title_bar.resizeEvent = lambda e: inner_container.setGeometry(0, 0, e.size().width(), e.size().height())
        
        layout = QHBoxLayout(title_bar)
        layout.setContentsMargins(15, 0, 15, 0)
        
        # Title and icon
        title_layout = QHBoxLayout()
        title_layout.setSpacing(4)
        
        icon = QLabel()
        icon_path = resource_path('myicon.ico')
        if os.path.exists(icon_path):
            app_icon = QIcon(icon_path)
            pixmap = app_icon.pixmap(20, 20)
            icon.setPixmap(pixmap)
            self.setWindowIcon(app_icon)
        
        title = QLabel("Genex ISDP Generator")
        title.setStyleSheet("color: white; font-weight: bold; font-size: 12px;")
        
        title_layout.addWidget(icon)
        title_layout.addWidget(title)
        title_layout.addStretch()
        
        # Window controls
        controls_layout = QHBoxLayout()
        controls_layout.setSpacing(4)
        
        # Minimize button
        minimize_button = QPushButton("(ˉ﹃ˉ)")
        minimize_button.setFixedSize(24, 24)
        minimize_button.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #FFFFFF;
                font-size: 20px;
                font-weight: bold;
                border: none;
                border-radius: 4px;
                margin: 0px;
                padding: 0px;
            }
            QPushButton:hover {
                background-color: rgba(255, 255, 255, 0.2);
            }
        """)
        minimize_button.clicked.connect(self.showMinimized)
        
        # Close button
        close_button = QPushButton("×")
        close_button.setFixedSize(24, 24)
        close_button.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #FFFFFF;
                font-size: 24px;
                font-weight: bold;
                border: none;
                border-radius: 4px;
                margin: 0px;
                padding: 0px;
            }
            QPushButton:hover {
                background-color: #FF0000;
            }
        """)
        close_button.clicked.connect(self.close)
        
        controls_layout.addWidget(minimize_button)
        controls_layout.addWidget(close_button)
        
        layout.addLayout(title_layout)
        layout.addStretch()
        layout.addLayout(controls_layout)
        
        return title_bar
        
    def create_file_section(self):
        file_frame = QFrame()
        file_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                                        stop:0 #008080, stop:1 #006666);
                border-radius: 6px;
            }
        """)
        
        layout = QHBoxLayout(file_frame)
        layout.setContentsMargins(4, 6, 4, 6)
        layout.setSpacing(4)
        
        self.file_input = QLineEdit()
        self.file_input.setPlaceholderText("Select ISDP EP file...")
        self.file_input.setReadOnly(True)
        self.file_input.setFixedHeight(26)
        self.file_input.setStyleSheet("""
            QLineEdit {
                background-color: rgba(255, 255, 255, 0.9);
                border: none;
                border-radius: 4px;
                color: #1e293b;
            }
        """)
        
        browse_button = QPushButton("Browse")
        browse_button.setFixedHeight(26)
        browse_button.setCursor(Qt.CursorShape.PointingHandCursor)
        browse_button.clicked.connect(self.select_file)
        browse_button.setStyleSheet("""
            QPushButton {
                background-color: white;
                color: #008080;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: rgb(222, 222, 222);
            }
        """)
        
        self.sheet_selector = QComboBox()
        self.sheet_selector.setPlaceholderText("Choose Sheet...")
        self.sheet_selector.setEnabled(False)
        self.sheet_selector.setFixedHeight(26)  # This matches the height of other elements
        self.sheet_selector.setStyleSheet("""
            QComboBox {
                background-color: white;
                border: none;
                border-radius: 4px;
                color: #1e293b;
                min-width: 150px;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: center right;
                width: 26px;
                height: 26px;  /* Match height to the combobox */
                border-left: none;
            }
            QComboBox::down-arrow {
                width: 16px;
                height: 16px;
            }
            QComboBox QAbstractItemView {
                background-color: transparent;
            }
        """)
        
        layout.addWidget(self.file_input, 1)
        layout.addWidget(browse_button)
        layout.addWidget(self.sheet_selector)
        
        return file_frame

    def create_filter_section(self):
        filter_frame = QFrame()
        filter_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 6px;
                border: 1px solid #e0e0e0;
            }
        """)
        
        layout = QHBoxLayout(filter_frame)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(4)
        
        filter_label = QLabel("Enter 4LRD:")
        filter_label.setStyleSheet("""
            color: #1e293b;
            font-weight: medium;
        """)
        
        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText("Enter 4LRD to filter...")
        self.filter_input.setFixedHeight(26)
        self.filter_input.setStyleSheet("""
            QLineEdit {
                background-color: #f8f9fa;
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                padding: 6px 12px;
            }
            QLineEdit:focus {
                border-color: #008080;
                background-color: white;
            }
        """)
        
        self.filter_button = QPushButton("Filter")
        self.filter_button.setFixedHeight(26)
        self.filter_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.filter_button.clicked.connect(self.filter_data)
        
        process_button = QPushButton("Start Parsing")
        process_button.setFixedHeight(26)
        process_button.setCursor(Qt.CursorShape.PointingHandCursor)
        process_button.clicked.connect(self.process_file)
        
        for button in [self.filter_button, process_button]:
            button.setStyleSheet("""
                QPushButton {
                    background-color: #008080;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 6px 16px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #006666;
                }
                QPushButton:pressed {
                    background-color: #004c4c;
                }
            """)
        
        layout.addWidget(filter_label)
        layout.addWidget(self.filter_input, 1)
        layout.addWidget(self.filter_button)
        layout.addWidget(process_button)
        
        return filter_frame
    
    def create_status_section(self):
        status_frame = QFrame()
        status_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 4px;
                border: 1px solid #e0e0e0;
            }
        """)
        
        layout = QVBoxLayout(status_frame)
        layout.setContentsMargins(1, 1, 1, 1)
        layout.setSpacing(1) 
        
        # Create two rows for status indicators
        top_row = QHBoxLayout()
        bottom_row = QHBoxLayout()
        
        top_row.setSpacing(1)
        bottom_row.setSpacing(1)
        
        # First row status items
        top_items = [
            ("eNodeB Name", "enodeb_name_label"),
            ("eNodeB ID", "enodeb_id_label"),
            ("TAC", "tac_label"),
            ("Sub Region", "subregion_label")
        ]
        
        # Second row status items
        bottom_items = [
            ("Lat/Long", "latlong_label"),
            ("Height", "height_label"),
            ("MTilt", "mtilt_label"),
            ("ETilt", "etilt_label")
        ]
        
        # Add items to top row
        for title, attr_name in top_items:
            status_widget = self.create_status_indicator(title)
            setattr(self, attr_name, status_widget['value'])
            top_row.addWidget(status_widget['container'])
        
        # Add items to bottom row
        for title, attr_name in bottom_items:
            status_widget = self.create_status_indicator(title)
            setattr(self, attr_name, status_widget['value'])
            bottom_row.addWidget(status_widget['container'])
        
        # Create containers for rows with reduced spacing
        top_container = QWidget()
        top_container.setLayout(top_row)
        bottom_container = QWidget()
        bottom_container.setLayout(bottom_row)
        
        # Add rows to main layout with reduced spacing
        layout.addWidget(top_container)
        layout.addSpacing(1)
        layout.addWidget(bottom_container)
        
        return status_frame

    def create_status_indicator(self, title):
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(1, 1, 1, 1)
        layout.setSpacing(2)
        
        title_label = QLabel(title)
        title_label.setStyleSheet("""
            color: #64748b;
            font-size: 11px;
            min-height: 14px;
            padding: 2px;
            border: 1px solid #008080;
        """)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
        title_label.setFixedWidth(75)
        
        value_label = QLabel()
        value_label.setStyleSheet("""
            color: #283747;
            font-size: 11px;
            font-weight: bold;
            min-height: 14px;
            padding: 1px 0;
            border: none;
            border-bottom: 2px solid #008080;
            background: transparent;
        """)
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
        value_label.setMinimumWidth(55)
        
        layout.addWidget(title_label)
        layout.addWidget(value_label)
        
        return {'container': container, 'value': value_label}

    def create_footer(self):
        footer = QWidget()
        footer_layout = QHBoxLayout(footer)
        footer_layout.setContentsMargins(0, 0, 0, 0)
        
        # Uptime label
        self.uptime_label = QLabel("00:00:00")
        self.uptime_label.setStyleSheet("""
            color: #64748b;
            font-size: 10px;
        """)
        
        # Version label with author credit
        version_label = QLabel("V4.2.5224 • Written in Python by Fadzli Abdullah")
        version_label.setStyleSheet("""
            color: #64748b;
            font-size: 10px;
        """)
        version_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Current date
        current_date = QDateTime.currentDateTime().toString("yyyy-MM-dd")
        date_label = QLabel(current_date)
        date_label.setStyleSheet("""
            color: #64748b;
            font-size: 10px;
        """)
        
        footer_layout.addWidget(self.uptime_label)
        footer_layout.addStretch()
        footer_layout.addWidget(version_label)
        footer_layout.addStretch()
        footer_layout.addWidget(date_label)
        
        footer.setStyleSheet("""
            QWidget {
                background-color: transparent;
                padding: 2px 0px;
            }
        """)
        
        return footer

    def format_worksheet(self, worksheet):
        """Format worksheet columns and cells"""
        try:
            # Set column widths
            standard_width = 15
            for column in worksheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                
                # Find maximum length in column
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                # Set width with some padding
                adjusted_width = max(standard_width, min(max_length + 2, 30))
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Set row height for header
            worksheet.row_dimensions[1].height = 30

            # Freeze top row
            worksheet.freeze_panes = 'A2'

        except Exception as e:
            print(f"Error in format_worksheet: {str(e)}")

    def apply_global_stylesheet(self):
        self.setStyleSheet("""
            QMainWindow {
                background: transparent;
            }
            QWidget {
                font-family: 'Roboto', 'Segoe UI', 'Arial';
                font-size: 11px;
                color: #1e293b;
            }
            QLineEdit {
                padding: 8px;
                border: 1px solid #e0e0e0;
                border-radius: 6px;
                background: #f8f9fa;
            }
            QLineEdit:focus {
                border: 1px solid #008080;
                background: white;
            }
            QPushButton {
                padding: 8px 16px;
                border-radius: 6px;
                border: none;
                background: #008080;
                color: white;
                font-weight: bold;
            }
            QPushButton:hover {
                background: #006666;
            }
            QPushButton:pressed {
                background: #004c4c;
            }
            QPushButton:disabled {
                background: #cccccc;
            }
            QProgressBar {
                border: none;
                background: #f0f0f0;
                border-radius: 2px;
            }
            QProgressBar::chunk {
                background: #008080;
                border-radius: 2px;
            }
            QComboBox {
                border: 1px solid #e0e0e0;
                border-radius: 6px;
                padding: 8px;
                background: white;
            }
            QMessageBox QPushButton {
                background-color: rgb(255,255,255);
                color: rgb(0, 0, 0);;
                padding: 6px 16px;
                border: none;
                border-radius: 4px;
                font-size: 11px;
                min-width: 80px;
            }
            QMessageBox QLabel {
                color: #1e293b;
                font-size: 11px;
            }
        """)
            
        # Window Movement Handlers
    def mousePressEvent(self, event):
        """Handle window dragging"""
        if event.button() == Qt.MouseButton.LeftButton:
            self._drag_pos = event.pos()

    def mouseMoveEvent(self, event):
        """Handle window movement"""
        if event.buttons() & Qt.MouseButton.LeftButton:
            if hasattr(self, '_drag_pos'):
                delta = event.pos() - self._drag_pos
                self.move(self.pos() + delta)

    def mouseReleaseEvent(self, event):
        """Handle end of window drag"""
        if hasattr(self, '_drag_pos'):
            del self._drag_pos

    # Utility Methods
    def center_window(self):
        """Center the window on the screen"""
        screen = QApplication.primaryScreen().geometry()
        window = self.geometry()
        x = (screen.width() - window.width()) // 2
        y = (screen.height() - window.height()) // 2
        self.move(x, y)

    def update_uptime(self):
        """Update the uptime display"""
        current_time = QDateTime.currentDateTime()
        elapsed = self.start_time.secsTo(current_time)
        hours = elapsed // 3600
        minutes = (elapsed % 3600) // 60
        seconds = elapsed % 60
        self.uptime_label.setText(f"{hours:02d}:{minutes:02d}:{seconds:02d}")

    def update_status(self, message):
        """Update the status message"""
        if hasattr(self, 'status_label'):
            self.status_label.setText(message)
            self.status_label.repaint()
            QApplication.processEvents()

    def update_status_values(self, **kwargs):
        for key, value in kwargs.items():
            if hasattr(self, f"{key}_label"):
                label = getattr(self, f"{key}_label")
                if isinstance(value, tuple):
                    text, color = value
                    label.setText(text)
                    if color:
                        label.setStyleSheet(f"""
                            color: {color};
                            font-size: 11px;
                            font-weight: bold;
                            padding: 1px 0;
                            border-bottom: 2px solid {color};  /* Only bottom border */
                            background: transparent;  /* Remove background */
                        """)
                else:
                    label.setText(str(value))

    def find_best_match_column(self, target_column, available_columns):
        """Find the best matching column using fuzzy matching"""
        best_match = None
        highest_ratio = 0

        for col in available_columns:
            ratio = fuzz.ratio(target_column.lower(), col.lower())
            if ratio > highest_ratio:
                highest_ratio = ratio
                best_match = col

        return best_match if highest_ratio > 70 else None

    def get_template_column_positions(self, worksheet):
        """Get the column positions from template header row"""
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value:
                    if cell.value in self.column_mapping:
                        source_col = self.column_mapping[cell.value][0]
                        self.column_mapping[cell.value] = (source_col, cell.column)
                    elif cell.value in self.static_values:
                        static_val = self.static_values[cell.value][0]
                        self.static_values[cell.value] = (static_val, cell.column)
                        
    def select_file(self):
        """File selection dialog and processing"""
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(
            self,
            "Select ISDP EP File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )

        if file_path:
            self.selected_file = file_path
            try:
                xl = pd.ExcelFile(file_path)
                sheet_names = xl.sheet_names

                self.sheet_selector.clear()
                self.sheet_selector.addItems(sheet_names)
                
                # Auto-select the first sheet if there's only one sheet
                if len(sheet_names) == 1:
                    self.sheet_selector.setCurrentIndex(0)
                    self.sheet_selector.setEnabled(False)
                else:
                    self.sheet_selector.setEnabled(True)
                
                self.file_input.setText(os.path.basename(file_path))
                
                # Clear previous status displays
                self.update_status_values(
                    enodeb_name="",
                    enodeb_id="",
                    tac="",
                    subregion="",
                    latlong="",
                    height="",
                    mtilt="",
                    etilt=""
                )
                
                # Process the file
                selected_sheet = self.sheet_selector.currentText()
                source_df = pd.read_excel(file_path, sheet_name=selected_sheet)
                
                # Find and process 4LRD column
                lrd_columns = [col for col in source_df.columns if '4LRD' in col or 'Site Name' in col]
                if lrd_columns:
                    lrd_column = lrd_columns[0]
                    unique_lrds = source_df[lrd_column].dropna().unique()
                    
                    if len(unique_lrds) == 1:
                        self.filtered_df = source_df
                        self.display_values_from_df(self.filtered_df)
                        self.filter_input.setText(str(unique_lrds[0]))
                        self.filter_input.setEnabled(False)
                        self.filter_button.setEnabled(False)
                        self.filter_button.setStyleSheet("""
                            QPushButton {
                                background-color: #94a3b8;
                                color: white;
                                border: none;
                                border-radius: 4px;
                                padding: 6px 16px;
                                font-weight: bold;
                            }
                        """)
                        self.update_status("✓ Single 4LRD found. Ready to process.")
                    else:
                        self.filter_input.setEnabled(True)
                        self.filter_input.clear()
                        self.filter_input.setFocus()
                        self.filtered_df = None
                        self.filter_button.setEnabled(True)
                        self.filter_button.setStyleSheet("""
                            QPushButton {
                                background-color: #008080;
                                color: white;
                                border: none;
                                border-radius: 4px;
                                padding: 6px 16px;
                                font-weight: bold;
                            }
                            QPushButton:hover {
                                background-color: #006666;
                            }
                        """)
                        self.update_status("✓ File loaded. Please enter 4LRD to filter data.")
                else:
                    self.update_status("⚠ Error: Could not find 4LRD or Site Name column!")

            except Exception as e:
                self.update_status(f"⚠ Error loading file: {str(e)}")
                print(f"Error details: {str(e)}")
        else:
            self.sheet_selector.clear()
            self.sheet_selector.addItem("No sheet available")
            self.sheet_selector.setEnabled(False)

    def filter_data(self):
        """Filter data based on 4LRD input"""
        if not self.selected_file:
            self.update_status("⚠ Error: Please select a file first!")
            return

        filter_text = self.filter_input.text().strip()
        if not filter_text:
            self.update_status("⚠ Please enter a 4LRD to filter!")
            return

        try:
            selected_sheet = self.sheet_selector.currentText()
            source_df = pd.read_excel(self.selected_file, sheet_name=selected_sheet)

            lrd_columns = [col for col in source_df.columns if '4LRD' in col or 'Site Name' in col]
            if lrd_columns:
                lrd_column = lrd_columns[0]
                self.filtered_df = source_df[source_df[lrd_column].str.contains(filter_text, case=False, na=False)]
                
                if self.filtered_df.empty:
                    self.update_status(f"⚠ No records found for 4LRD: {filter_text}")
                    return
                
                self.display_values_from_df(self.filtered_df)
                self.update_status(f"✓ Found {len(self.filtered_df)} records for 4LRD: {filter_text}")
            else:
                self.update_status("⚠ Error: Could not find 4LRD or Site Name column!")

        except Exception as e:
            self.update_status(f"⚠ Error during filtering: {str(e)}")
            print(f"Error details: {str(e)}")
            
    def check_consistency(self, df, column_keywords, column_type=""):
        """Generic consistency checker for columns"""
        try:
            matching_columns = [col for col in df.columns 
                            if any(keyword.lower() in col.lower() 
                                    for keyword in column_keywords)]
            
            if matching_columns:
                target_col = matching_columns[0]
                values = df[target_col].dropna()
                
                is_consistent = len(values) > 0
                status = "OKAY" if is_consistent else "NOT OKAY"
                color = "#22c55e" if is_consistent else "#ef4444"
                
                return status, color
            return "NOT OKAY", "#ef4444"
            
        except Exception as e:
            print(f"Error checking {column_type} consistency: {str(e)}")
            return "NOT OKAY", "#ef4444"

    def display_values_from_df(self, df):
        try:
            columns = {col.lower(): col for col in df.columns}
            
            # Column mappings
            field_mappings = {
                'enodeb_name': ['enodeb name', 'nodeb name', 'enodebname', 'nodebname', 'enb name', 'enbname'],
                'enodeb_id': ['enodebid', 'enodeb id', 'nodebid'],
                'tac': ['tac', 'tracking area code'],
                'subregion': ['kv border', 'sub region', 'subregion', 'sub-region']
            }
            
            # Extract values
            values = {}
            for field, keywords in field_mappings.items():
                col_name = next((columns[key] for key in keywords if key in columns), None)
                if col_name and not df[col_name].empty:
                    val = df[col_name].iloc[0]
                    if pd.notna(val):
                        values[field] = str(val).strip()
                    else:
                        values[field] = ""
            
            # Check consistencies
            latlong_status, latlong_color = self.check_consistency(
                df, ['long', 'lng', 'longitude', 'lat', 'latitude'], "latlong")
            height_status, height_color = self.check_consistency(
                df, ['height', 'ant height', 'antenna height'], "height")
            mtilt_status, mtilt_color = self.check_consistency(
                df, ['plan mechanical tilt', 'mechanical tilt', 'mtilt'], "mtilt")
            etilt_status, etilt_color = self.check_consistency(
                df, ['plan ret', 'electrical tilt', 'etilt'], "etilt")


            # Update status indicators with priority for KV BORDER
            subregion_value = (
                values.get('kv border') or  # Try KV BORDER first
                values.get('sub region') or  # Then try Sub Region
                values.get('subregion') or   # Then other variations
                values.get('sub-region') or 
                ''
            )

            # Update status indicators
            self.update_status_values(
                enodeb_name=values.get('enodeb_name', ''),
                enodeb_id=values.get('enodeb_id', ''),
                tac=values.get('tac', ''),
                subregion=subregion_value,  # Use the prioritized subregion value
                latlong=(latlong_status, latlong_color),
                height=(height_status, height_color),
                mtilt=(mtilt_status, mtilt_color),
                etilt=(etilt_status, etilt_color)
            )

        except Exception as e:
                print(f"Error in display_values_from_df: {str(e)}")

    def get_subregion_column(self, df):
        """Helper method to determine which column to use for subregion information"""
        possible_columns = ['KV BORDER', 'Sub Region', 'SubRegion', 'Sub-Region']
        for col in possible_columns:
            if col in df.columns:
                return col
        return None

    def process_file(self):
        """Process the file with filtered data"""
        if not self.selected_file:
            self.update_status("⚠ Error: Please select a file first!")
            return

        # If filtered_df is None, try to set it using the current sheet selection
        if self.filtered_df is None or self.filtered_df.empty:
            selected_sheet = self.sheet_selector.currentText()
            if selected_sheet and selected_sheet != "No sheet available":
                try:
                    source_df = pd.read_excel(self.selected_file, sheet_name=selected_sheet)
                    
                    # Check if we need to filter or can process directly
                    lrd_columns = [col for col in source_df.columns if '4LRD' in col or 'Site Name' in col]
                    if lrd_columns:
                        lrd_column = lrd_columns[0]
                        filter_text = self.filter_input.text().strip()
                        
                        if filter_text:
                            # Apply filter if text is provided
                            self.filtered_df = source_df[source_df[lrd_column].str.contains(filter_text, case=False, na=False)]
                        else:
                            # If no filter text but only one unique LRD, use the whole dataframe
                            unique_lrds = source_df[lrd_column].dropna().unique()
                            if len(unique_lrds) == 1:
                                self.filtered_df = source_df
                            else:
                                self.update_status("⚠ Please enter a 4LRD to filter!")
                                return
                    else:
                        self.update_status("⚠ Error: Could not find 4LRD or Site Name column!")
                        return
                except Exception as e:
                    self.update_status(f"⚠ Error processing file: {str(e)}")
                    print(f"Error details: {str(e)}")
                    return
            else:
                self.update_status("⚠ No data to process!")
                return

        if self.filtered_df.empty:
            self.update_status("⚠ No data to process!")
            return

        try:
            self.update_status("Processing data...")
            source_df = self.filtered_df.reset_index(drop=True)
            
            # Get the 4LRD value for the output filename
            lrd_columns = [col for col in source_df.columns if '4LRD' in col or 'Site Name' in col]
            if lrd_columns:
                lrd_value = str(source_df[lrd_columns[0]].iloc[0]).strip()
            else:
                lrd_value = "output"

            # Template file paths
            possible_template_paths = [
                os.path.join(os.path.dirname(self.selected_file), "LTEEngineeringParameterTemplate.xlsx"),
                os.path.join(os.path.expanduser("~/Desktop"), "LTEEngineeringParameterTemplate.xlsx"),
                os.path.join(os.getcwd(), "LTEEngineeringParameterTemplate.xlsx"),
                "LTEEngineeringParameterTemplate.xlsx"
            ]
            
            template_path = next((path for path in possible_template_paths if os.path.exists(path)), None)
            
            if template_path is None:
                self.update_status("⚠ Error: Template file not found!")
                return
            
            # Process Excel workbook
            wb = openpyxl.load_workbook(template_path)
            # Get the first worksheet if 'LTE' is not found
            if 'LTE' in wb.sheetnames:
                ws = wb['LTE']
            else:
                ws = wb.active  # Get the active worksheet
                
            self.get_template_column_positions(ws)

            # Clear existing content
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None

            # Set styles
            output_font = Font(name='Tahoma', size=11)
            header_alignment = Alignment(horizontal='center', vertical='center')

            # Find the Electrical Downtilt column in source dataframe
            etilt_column = self.find_best_match_column('Electrical Downtilt', source_df.columns)

            # Process rows
            for idx, source_row in enumerate(source_df.itertuples(), start=2):
                # Map columns
                for template_col, (source_col, col_idx) in self.column_mapping.items():
                    if col_idx:
                        cell = ws.cell(row=idx, column=col_idx)
                        
                        # Special handling for Electrical Downtilt - place default value of 2 when value is 0 or blank
                        if template_col == 'Electrical Downtilt':
                            # Set default value of 2 first
                            cell.value = 2
                            
                            # Only override if we find a non-zero, non-blank value
                            best_match = self.find_best_match_column(source_col, source_df.columns)
                            if best_match:
                                value = source_row[source_df.columns.get_loc(best_match) + 1]
                                # Only use the value if it's not 0, blank, or null
                                if not pd.isna(value) and value != '' and value != 0 and value is not None:
                                    cell.value = value
                        else:
                            best_match = self.find_best_match_column(source_col, source_df.columns)
                            if best_match:
                                cell.value = source_row[source_df.columns.get_loc(best_match) + 1]
                            
                        cell.font = output_font
                        cell.alignment = header_alignment

                # Apply static values
                for template_col, (static_value, col_idx) in self.static_values.items():
                    if col_idx:
                        cell = ws.cell(row=idx, column=col_idx)
                        cell.value = static_value
                        cell.font = output_font
                        cell.alignment = header_alignment

            # Format columns
            self.format_worksheet(ws)
            
            # Create output filename with timestamp
            timestamp = QDateTime.currentDateTime().toString("yyyyMMdd_HHmmss")
            output_filename = f"ISDP_{lrd_value}_{timestamp}.xlsx"
            output_path = os.path.join(self.output_folder, output_filename)
            
            # Ensure the output directory exists
            os.makedirs(self.output_folder, exist_ok=True)
            
            # Save the workbook
            wb.save(output_path)
            
            # Play sound notification
            winsound.MessageBeep(winsound.MB_OK)
            
            # Open the file
            if os.path.exists(output_path):
                os.startfile(output_path)
                self.update_status(f"✓ Process completed successfully! Processed {len(source_df)} records.")
            else:
                self.update_status("⚠ Error: Failed to create output file!")

        except Exception as e:
            self.update_status(f"⚠ Error: {str(e)}")
            print(f"Error details: {str(e)}")
            
if __name__ == '__main__':
    try:
        app = QApplication(sys.argv)
        app.setStyle('Fusion')

        # Set application icon
        icon_path = resource_path('myicon.ico')
        if os.path.exists(icon_path):
            app.setWindowIcon(QIcon(icon_path))
        
        # Set application-wide color palette
        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window, QColor("#ffffff"))
        palette.setColor(QPalette.ColorRole.WindowText, QColor("#1e293b"))
        palette.setColor(QPalette.ColorRole.Base, QColor("#ffffff"))
        palette.setColor(QPalette.ColorRole.AlternateBase, QColor("#f8f9fa"))
        palette.setColor(QPalette.ColorRole.ToolTipBase, QColor("#ffffff"))
        palette.setColor(QPalette.ColorRole.ToolTipText, QColor("#1e293b"))
        palette.setColor(QPalette.ColorRole.Text, QColor("#1e293b"))
        palette.setColor(QPalette.ColorRole.Button, QColor("#ffffff"))
        palette.setColor(QPalette.ColorRole.ButtonText, QColor("#1e293b"))
        palette.setColor(QPalette.ColorRole.Link, QColor("#008080"))
        palette.setColor(QPalette.ColorRole.Highlight, QColor("#008080"))
        palette.setColor(QPalette.ColorRole.HighlightedText, QColor("#ffffff"))
        
        app.setPalette(palette)
        
        window = ModernExcelProcessor()
        window.show()
        sys.exit(app.exec())
        
    except Exception as e:
        print(f"Application error: {str(e)}")
        import traceback
        traceback.print_exc()
