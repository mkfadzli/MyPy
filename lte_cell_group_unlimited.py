import sys
import os
import subprocess
import re
from PyQt6.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, 
                           QPushButton, QLabel, QLineEdit, QTextEdit, QWidget, 
                           QFileDialog, QMessageBox, QFrame)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QIcon, QPalette, QColor
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import traceback


def sanitize_filename(filename):
    """Sanitize filename by removing invalid characters"""
    # Remove invalid characters for Windows filenames
    invalid_chars = r'[<>:"/\\|?*]'
    sanitized = re.sub(invalid_chars, '_', filename)
    # Remove leading/trailing spaces and dots
    sanitized = sanitized.strip('. ')
    # Limit length to avoid filesystem issues
    if len(sanitized) > 100:
        sanitized = sanitized[:100]
    return sanitized if sanitized else "Cell_Group"


class ProcessingThread(QThread):
    """Thread for processing Excel data to avoid GUI freezing"""
    finished = pyqtSignal(bool, str, str)  # Added third parameter for output file path
    
    def __init__(self, genexep_file, lrd_list, cell_group_name, template_dir, output_dir):
        super().__init__()
        self.genexep_file = genexep_file
        self.lrd_list = lrd_list
        self.cell_group_name = cell_group_name
        self.template_dir = template_dir
        self.output_dir = output_dir
    
    def run(self):
        try:
            # Define paths
            template_source = os.path.join(self.template_dir, "Export_cell_group_template.xlsx")
            
            # Create output filename based on Cell Group Name
            sanitized_name = sanitize_filename(self.cell_group_name)
            output_filename = f"{sanitized_name}_Cell_Group.xlsx"
            # output_path will be constructed later using self.output_dir
            
            # Check if template source exists
            if not os.path.exists(template_source):
                self.finished.emit(False, f"Template file not found: {template_source}", "")
                return

            # Load the PRS Object Tree workbook (header at row 3 -> header=2)
            # We expect the column header 'Cell Name' to be at column H (header row is row 3)
            prs_df = pd.read_excel(self.genexep_file, header=2, dtype=str)

            # Find important columns (case-insensitive). We require 'Cell Name'.
            cols_map = {c.strip().lower(): c for c in prs_df.columns}

            def find_col(candidates):
                for cand in candidates:
                    key = cand.strip().lower()
                    if key in cols_map:
                        return cols_map[key]
                # fallback: contains match
                for cand in candidates:
                    key = cand.strip().lower()
                    for k, orig in cols_map.items():
                        if key in k:
                            return orig
                return None

            cell_name_col = find_col(['cell name', 'cellname'])
            if not cell_name_col:
                self.finished.emit(False, "Could not find 'Cell Name' column in PRS Object Tree (expected header at row 3, column H).", "")
                return

            cell_id_col = find_col(['cell id', 'cellid', 'cid'])
            enodeb_id_col = find_col(['enodeb id', 'enodebid', 'enodeb id'])
            enodeb_name_col = find_col(['enodeb name', 'enodebname', 'e nodeb name', 'enodeb'])

            # Create 4LRD from first 4 characters of Cell Name (uppercase)
            prs_df['4LRD'] = prs_df[cell_name_col].astype(str).str.upper().str[:4]

            # Filter rows where 4LRD matches any of the provided LRDs
            lrd_set = set([l.upper() for l in self.lrd_list])
            filtered_data = prs_df[prs_df['4LRD'].isin(lrd_set)].copy()

            if filtered_data.empty:
                self.finished.emit(False, "No matching data found for the provided 4LRDs in PRS Object Tree.", "")
                return

            # Prepare rows list with the columns we will write to the template
            rows = []
            for _, r in filtered_data.iterrows():
                rows.append({
                    'cell_id': r.get(cell_id_col, '') if cell_id_col else '',
                    'cell_name': r.get(cell_name_col, ''),
                    'enodeb_id': r.get(enodeb_id_col, '') if enodeb_id_col else '',
                    'enodeb_name': r.get(enodeb_name_col, '') if enodeb_name_col else ''
                })

            # Split rows into chunks respecting eNodeB Name groups and max rows limit
            MAX_ROWS = 10000
            chunks = []
            current = []
            for r in rows:
                current.append(r)
                if len(current) > MAX_ROWS:
                    # Determine the last group's name in current chunk
                    last_group = current[-1].get('enodeb_name', '')
                    # Find the first index of this last_group in current
                    split_idx = None
                    for idx, item in enumerate(current):
                        if item.get('enodeb_name', '') == last_group:
                            split_idx = idx
                            break

                    if split_idx is None or split_idx == 0:
                        # The group itself is larger than MAX_ROWS or no earlier boundary: finalize whole current
                        chunks.append(current)
                        current = []
                    else:
                        # Finalize up to split_idx (exclude the group that caused overflow)
                        chunks.append(current[:split_idx])
                        # Start new current with the overflowing group
                        current = current[split_idx:]

            if current:
                chunks.append(current)

            # Write each chunk into its own workbook (use part suffix if more than one)
            output_paths = []
            total_written = 0
            part = 1
            for chunk in chunks:
                wb = load_workbook(template_source)
                ws = wb.active

                # Clear existing data
                for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row_cells:
                        cell.value = None

                center_alignment = Alignment(horizontal='center', vertical='center')
                row_num = 2
                for item in chunk:
                    ws.cell(row=row_num, column=1, value=self.cell_group_name).alignment = center_alignment
                    ws.cell(row=row_num, column=2, value=item.get('cell_id', '')).alignment = center_alignment
                    ws.cell(row=row_num, column=3, value=item.get('cell_name', '')).alignment = center_alignment
                    ws.cell(row=row_num, column=4, value=item.get('enodeb_id', '')).alignment = center_alignment
                    ws.cell(row=row_num, column=5, value=item.get('enodeb_name', '')).alignment = center_alignment
                    ws.cell(row=row_num, column=6, value='admin').alignment = center_alignment
                    row_num += 1

                # Verification check: Remove rows where 5th character of eNodeB Name is "B"
                rows_to_delete = []
                for r_idx in range(2, ws.max_row + 1):
                    enodeb_name_cell = ws.cell(row=r_idx, column=5)
                    enodeb_name = str(enodeb_name_cell.value) if enodeb_name_cell.value else ""
                    if len(enodeb_name) >= 5 and enodeb_name[4] == "B":
                        rows_to_delete.append(r_idx)
                for r_idx in reversed(rows_to_delete):
                    ws.delete_rows(r_idx)

                # Auto-fit columns
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value is not None and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Save file
                if len(chunks) == 1:
                    out_name = output_filename
                else:
                    out_name = f"{sanitized_name}_Cell_Group_part{part}.xlsx"
                out_path = os.path.join(self.template_dir, out_name)
                wb.save(out_path)
                wb.close()

                output_paths.append(out_path)
                total_written += (ws.max_row - 1 if ws.max_row > 1 else 0)
                part += 1

            # Compose success message
            if len(output_paths) == 1:
                success_msg = f"LTE Cell Group template created!\n\n{total_written} records processed and saved to:\n{os.path.basename(output_paths[0])}"
            else:
                files_list = '\n'.join([os.path.basename(p) for p in output_paths])
                success_msg = f"LTE Cell Group templates created!\n\n{total_written} records processed and saved to:\n{files_list}\n\nFiles were split to respect the {MAX_ROWS} row limit without splitting the same eNodeB Name across files."
            # Emit success with first file path for opening
            self.finished.emit(True, success_msg, output_paths[0] if output_paths else "")
            
            
            
        except Exception as e:
            error_msg = f"Error processing data: {str(e)}\n{traceback.format_exc()}"
            self.finished.emit(False, error_msg, "")


class GenexEPApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.genexep_file_path = ""
        self.template_dir = r"C:\Users\mWX1318105\Desktop\LTE_Cell_Group_Creator"
        self.output_dir = self.template_dir  # default output directory
        self.output_file_path = ""  # Store the path of the generated file
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle("LTE Cell Group Creator")
        
        # Set window icon (assumes icon.ico is in the same folder)
        icon_path = os.path.join(os.path.dirname(__file__), "icon.ico")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        
        self.setFixedSize(440, 580)  # Increased height for better spacing
        
        # Set modern style
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QPushButton {
                background-color: #4CAF50;
                border: none;
                color: white;
                padding: 4px 10px;
                text-align: center;
                font-size: 12px;
                font-weight: bold;
                border-radius: 4px;
                min-height: 18px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
            QLineEdit, QTextEdit {
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 2px;
                font-size: 11px;
                background-color: white;
                color: #333333;
            }
            QLineEdit:focus, QTextEdit:focus {
                border-color: #4CAF50;
            }
            QLabel {
                color: #333;
                font-size: 11px;
                font-weight: bold;
            }
            .status-label {
                color: #2196F3;
                font-size: 10px;
                font-weight: normal;
                font-style: italic;
            }
            .counter-label {
                color: #666666;
                font-size: 10px;
                font-weight: normal;
                font-style: italic;
            }
            QLabel[objectName="footer"] {
                color: #888888;
                font-size: 10px;
                font-weight: normal;
            }
        """)
        
        # Create central widget and main layout
        central_widget = QWidget()
        central_widget.setObjectName("central_widget")
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 12, 15, 12)
        
        # Load PRS Object Tree button
        self.load_button = QPushButton("📁 Select PRS Object File")
        self.load_button.setObjectName("load_button")
        self.load_button.clicked.connect(self.load_genexep_file)
        main_layout.addWidget(self.load_button)
        
        # Status label for file upload completion
        self.status_label = QLabel("")
        self.status_label.setObjectName("status_label")
        self.status_label.setStyleSheet("color: #000000; font-size: 10px; font-weight: normal; font-style: italic;")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_label.hide()  # Initially hidden
        main_layout.addWidget(self.status_label)
        
        # 4LRD input section
        lrd_container = QWidget()
        lrd_layout = QVBoxLayout(lrd_container)
        lrd_layout.setContentsMargins(0, 0, 0, 0)
        lrd_layout.setSpacing(6)
        
        lrd_label = QLabel("🏢 4LRD 🤳")
        lrd_sublabel = QLabel("Enter 4LRD, 5LRD or 6LRD (unlimited)")
        lrd_sublabel.setStyleSheet("color: #64748b; font-size: 11px; font-weight: 400; margin-bottom: 0px;")
        
        lrd_layout.addWidget(lrd_label)
        lrd_layout.addWidget(lrd_sublabel)
        
        self.lrd_input = QTextEdit()
        # Make the input field taller so roughly 4 rows are visible
        self.lrd_input.setFixedHeight(140)
        self.lrd_input.setPlaceholderText("Paste Site Name or eNodeB Name here...")
        self.lrd_input.textChanged.connect(self.format_lrd_input)
        lrd_layout.addWidget(self.lrd_input)
        
        # 4LRD Counter label - positioned right below the input field
        self.lrd_counter_label = QLabel("0 4LRDs entered")
        self.lrd_counter_label.setObjectName("counter_label")
        self.lrd_counter_label.setStyleSheet("color: #666666; font-size: 10px; font-weight: normal; font-style: italic;")
        self.lrd_counter_label.setAlignment(Qt.AlignmentFlag.AlignLeft)
        lrd_layout.addWidget(self.lrd_counter_label)
        
        main_layout.addWidget(lrd_container)
        
        # Cell Group Name section
        cell_container = QWidget()
        cell_layout = QVBoxLayout(cell_container)
        cell_layout.setContentsMargins(0, 0, 0, 0)
        cell_layout.setSpacing(6)
        
        cell_group_label = QLabel("📋 Cell Group Name 🐱‍👓")
        cell_layout.addWidget(cell_group_label)
        
        self.cell_group_input = QLineEdit()
        self.cell_group_input.setPlaceholderText("Enter Cell Group Name")
        cell_layout.addWidget(self.cell_group_input)
        
        main_layout.addWidget(cell_container)

        # Output folder selector
        out_container = QWidget()
        out_layout = QHBoxLayout(out_container)
        out_layout.setContentsMargins(0, 0, 0, 0)
        out_layout.setSpacing(8)

        self.output_button = QPushButton("📂 Select Output Folder")
        self.output_button.clicked.connect(self.select_output_folder)
        out_layout.addWidget(self.output_button)

        self.output_label = QLabel(self.template_dir)
        self.output_label.setToolTip(self.template_dir)
        self.output_label.setStyleSheet("color: #666666; font-size: 10px;")
        out_layout.addWidget(self.output_label)

        main_layout.addWidget(out_container)
        
        # Generate button
        self.generate_button = QPushButton("⚡ Create Template")
        self.generate_button.setObjectName("generate_button")
        self.generate_button.clicked.connect(self.generate_template)
        self.generate_button.setEnabled(False)
        main_layout.addWidget(self.generate_button)
        
        # Smaller fixed spacer to keep consistent padding above footer
        main_layout.addSpacing(10)
        
        # Footer
        footer_label = QLabel("V3.0. Created by Fadzli Abdullah")
        footer_label.setObjectName("footer")
        footer_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(footer_label)
        
    def load_genexep_file(self):
        """Open file dialog to select GenexEP Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select PRS Object Tree Excel File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            self.genexep_file_path = file_path
            filename = os.path.basename(file_path)
            # Truncate long filenames for better display
            if len(filename) > 25:
                filename = filename[:22] + "..."
            self.load_button.setText(f"✅ {filename}")
            
            # Show completion status
            self.status_label.setText("✓ File loaded successfully")
            self.status_label.show()
            
            self.check_generate_button()

    def select_output_folder(self):
        """Allow user to select output folder for generated files"""
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder", self.output_dir)
        if folder:
            self.output_dir = folder
            # Update label and tooltip
            display = folder
            if len(display) > 60:
                display = '...' + display[-57:]
            self.output_label.setText(display)
            self.output_label.setToolTip(folder)
            self.check_generate_button()
    
    def format_lrd_input(self):
        """Auto-format 4LRD input with proper comma separation (accepts 6LRD and truncates to 4LRD) and removes duplicates"""
        # Get current cursor position to restore it later
        cursor = self.lrd_input.textCursor()
        position = cursor.position()
        
        # Get current text
        current_text = self.lrd_input.toPlainText()
        
        # Don't format if text is empty or only whitespace
        if not current_text.strip():
            # Update counter for empty input
            self.lrd_counter_label.setText("0 4LRDs entered")
            # Always check generate button state
            self.check_generate_button()
            return
        
        # Split by various delimiters and clean up
        import re
        # Split by comma, semicolon, space, newline, tab
        items = re.split(r'[,;\s\n\t]+', current_text)
        
        # Clean up items: remove empty strings, strip whitespace, convert to uppercase
        cleaned_items = []
        seen = set()  # To track duplicates
        
        for item in items:
            item = item.strip().upper()
            if item and len(item) <= 6:  # Accept items up to 6 characters (6LRD)
                # Truncate to 4 characters if longer than 4 (convert 6LRD to 4LRD)
                truncated_item = item[:4]
                # Only add if not already seen (removes duplicates while preserving order)
                if truncated_item not in seen:
                    cleaned_items.append(truncated_item)
                    seen.add(truncated_item)
        
        # Update the counter label
        count = len(cleaned_items)
        if count == 0:
            self.lrd_counter_label.setText("0 4LRDs entered")
        elif count == 1:
            self.lrd_counter_label.setText("1 4LRD entered")
        else:
            self.lrd_counter_label.setText(f"{count} 4LRDs entered")
        
        # Join with commas and space for readability
        formatted_text = ', '.join(cleaned_items)
        
        # Only update if the text actually changed to avoid infinite recursion
        if formatted_text != current_text:
            # Temporarily disconnect the signal to avoid recursion
            self.lrd_input.textChanged.disconnect()
            
            # Update the text
            self.lrd_input.setPlainText(formatted_text)
            
            # Restore cursor position (adjust for new text length)
            new_position = min(position, len(formatted_text))
            cursor.setPosition(new_position)
            self.lrd_input.setTextCursor(cursor)
            
            # Reconnect the signal
            self.lrd_input.textChanged.connect(self.format_lrd_input)
        
        # Always check if generate button should be enabled after any input change
        self.check_generate_button()
    
    def check_generate_button(self):
        """Enable generate button if all required fields are filled"""
        has_file = bool(self.genexep_file_path)
        has_lrd = bool(self.lrd_input.toPlainText().strip())
        has_cell_group = bool(self.cell_group_input.text().strip())
        
        self.generate_button.setEnabled(has_file and has_lrd and has_cell_group)
    
    def show_sharp_message_box(self, icon, title, message):
        """Show message box with crisp, sharp fonts"""
        msg = QMessageBox(self)
        msg.setIcon(icon)
        msg.setWindowTitle(title)
        msg.setText(message)
        
        # Set a clear, sharp font for the message box
        font = QFont("Tahoma", 9)  # Use system font with specific size
        font.setHintingPreference(QFont.HintingPreference.PreferFullHinting)
        font.setStyleStrategy(QFont.StyleStrategy.PreferAntialias)
        msg.setFont(font)
        
        # Apply custom stylesheet for better text rendering and improved color contrast
        msg.setStyleSheet("""
            QMessageBox {
                font-family: "Segoe UI", "Arial", sans-serif;
                font-size: 9pt;
                font-weight: normal;
                background-color: #ffffff;
            }
            QMessageBox QLabel {
                font-family: "Segoe UI", "Arial", sans-serif;
                font-size: 9pt;
                font-weight: normal;
                color: #2c2c2c;
                background-color: transparent;
            }
            QPushButton {
                font-family: "Segoe UI", "Arial", sans-serif;
                font-size: 9pt;
                font-weight: normal;
                min-width: 75px;
                min-height: 23px;
                padding: 4px 12px;
                background-color: #f0f0f0;
                border: 1px solid #adadad;
                border-radius: 3px;
                color: #2c2c2c;
            }
            QPushButton:hover {
                background-color: #e5f1fb;
                border-color: #0078d4;
            }
            QPushButton:pressed {
                background-color: #cde8ff;
                border-color: #005a9e;
            }
        """)
        
        return msg.exec()
    
    def generate_template(self):
        """Process the data and generate the template"""
        # Validate template directory exists
        if not os.path.exists(self.template_dir):
            self.show_sharp_message_box(
                QMessageBox.Icon.Critical, 
                "Fadzli Abdullah", 
                f"Template directory not found: {self.template_dir}"
            )
            return
        
        # Validate template file exists
        template_source = os.path.join(self.template_dir, "Export_cell_group_template.xlsx")
        if not os.path.exists(template_source):
            self.show_sharp_message_box(
                QMessageBox.Icon.Critical, 
                "Fadzli Abdullah", 
                f"Template file not found: {template_source}"
            )
            return
        
        # Parse 4LRD input
        lrd_text = self.lrd_input.toPlainText().strip()
        lrd_list = [lrd.strip().upper() for lrd in lrd_text.split(',') if lrd.strip()]
        
        # Removed the 1000 limit check
        
        if not lrd_list:
            self.show_sharp_message_box(
                QMessageBox.Icon.Warning, 
                "Fadzli Abdullah", 
                "Please enter at least one 4LRD."
            )
            return
        
        cell_group_name = self.cell_group_input.text().strip()
        
        # Disable generate button during processing
        self.generate_button.setEnabled(False)
        self.generate_button.setText("🔄 Processing...")
        
        # Start processing in separate thread
        self.processing_thread = ProcessingThread(
            self.genexep_file_path, lrd_list, cell_group_name, self.template_dir, self.output_dir
        )
        self.processing_thread.finished.connect(self.on_processing_finished)
        self.processing_thread.start()
    
    def on_processing_finished(self, success, message, output_path):
        """Handle processing completion"""
        self.generate_button.setText("⚡ Create Template")
        self.check_generate_button()
        
        if success:
            self.output_file_path = output_path
            
            # Show success message with sharp fonts
            self.show_sharp_message_box(
                QMessageBox.Icon.Information, 
                "Fadzli Abdullah", 
                message
            )
            
            # Open the updated template file
            try:
                if sys.platform.startswith('win'):
                    os.startfile(self.output_file_path)
                elif sys.platform.startswith('darwin'):  # macOS
                    subprocess.run(['open', self.output_file_path])
                else:  # Linux
                    subprocess.run(['xdg-open', self.output_file_path])
            except Exception as e:
                self.show_sharp_message_box(
                    QMessageBox.Icon.Warning, 
                    "Fadzli Abdullah", 
                    f"File processed successfully but couldn't open automatically: {str(e)}"
                )
        else:
            self.show_sharp_message_box(
                QMessageBox.Icon.Critical, 
                "Fadzli Abdullah", 
                message
            )
    
    def keyPressEvent(self, event):
        """Handle key press events for better UX"""
        # Enable generate button checking when text changes
        if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter, Qt.Key.Key_Tab):
            self.check_generate_button()
        super().keyPressEvent(event)


def main():
    # Enable High DPI scaling for sharp fonts
    if hasattr(Qt, 'AA_EnableHighDpiScaling'):
        QApplication.setAttribute(Qt.ApplicationAttribute.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
        QApplication.setAttribute(Qt.ApplicationAttribute.AA_UseHighDpiPixmaps, True)
    
    app = QApplication(sys.argv)
    
    # Set application properties
    app.setApplicationName("LTE Cell Group Creator")
    app.setApplicationVersion("1.0.2225")
    app.setOrganizationName("Huawei Technologies")
    
    # Set a default font with better rendering for the entire application
    default_font = QFont("Tahoma", 9)
    default_font.setHintingPreference(QFont.HintingPreference.PreferFullHinting)
    default_font.setStyleStrategy(QFont.StyleStrategy.PreferAntialias)
    app.setFont(default_font)
    
    # Create and show the main window
    window = GenexEPApp()
    window.show()
    
    # Connect text change events to check generate button
    window.cell_group_input.textChanged.connect(window.check_generate_button)
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()