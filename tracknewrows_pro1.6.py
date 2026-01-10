import threading
import queue
import time
import csv
from pathlib import Path
import traceback
import sys
from datetime import datetime
import os
import subprocess
import platform

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTextEdit, QFileDialog,
    QMessageBox, QProgressBar, QCheckBox, QFrame, QGroupBox
)
from PyQt6.QtCore import Qt, QTimer, pyqtSignal, QObject
from PyQt6.QtGui import QFont, QIcon

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font as ExcelFont, PatternFill, Alignment
except Exception as e:
    print("Missing dependency 'openpyxl'. Install with: pip install openpyxl", file=sys.stderr)
    raise

# -------------------------
# Core streaming functions
# -------------------------
def normalize(v):
    """Normalize values for comparison - handles None, strips whitespace"""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()

def read_header_and_index_map_xlsx(path: Path, sheet_name=None):
    """Read Excel header and create column name to index mapping"""
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    it = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    header = next(it)
    header = [h if h is not None else "" for h in header]
    name_to_idx = {str(name).strip(): idx for idx, name in enumerate(header)}
    wb.close()
    return header, name_to_idx

def build_key_set_from_xlsx(path: Path, key_cols: list, sheet_name=None, progress_callback=None):
    """
    Build a set of unique keys from Excel file.
    CRITICAL: Uses composite key from multiple columns to ensure uniqueness per row.
    """
    header, name_to_idx = read_header_and_index_map_xlsx(path, sheet_name)
    key_indexes = []
    
    # Validate all key columns exist
    for k in key_cols:
        if k not in name_to_idx:
            raise ValueError(f"Key column '{k}' not found in '{path.name}' header. Available columns: {list(name_to_idx.keys())}")
        key_indexes.append(name_to_idx[k])

    s = set()
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    first = True
    scanned = 0
    
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        scanned += 1
        # Build composite key from all specified columns
        key = tuple(normalize(row[i]) for i in key_indexes)
        s.add(key)
        
        if progress_callback and (scanned % 50000 == 0):
            progress_callback(f"Scanned {scanned:,} rows from old file...")
    
    wb.close()
    return s

def build_row_data_from_xlsx(path: Path, key_cols: list, sheet_name=None, progress_callback=None):
    """
    Build a dictionary mapping keys to full row data for change detection.
    Returns: (row_data_dict, header)
    """
    header, name_to_idx = read_header_and_index_map_xlsx(path, sheet_name)
    key_indexes = []
    
    # Validate all key columns exist
    for k in key_cols:
        if k not in name_to_idx:
            raise ValueError(f"Key column '{k}' not found in '{path.name}' header. Available columns: {list(name_to_idx.keys())}")
        key_indexes.append(name_to_idx[k])

    row_data = {}
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    first = True
    scanned = 0
    
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        scanned += 1
        # Build composite key from all specified columns
        key = tuple(normalize(row[i]) for i in key_indexes)
        row_data[key] = {header[i]: row[i] for i in range(len(row))}
        
        if progress_callback and (scanned % 50000 == 0):
            progress_callback(f"Scanned {scanned:,} rows from old file...")
    
    wb.close()
    return row_data, header

def find_all_changes_xlsx(old_path: Path, new_path: Path, key_cols: list, out_path: Path,
                          sheet_name=None, out_xlsx=False, status_callback=None, progress_callback=None):
    """
    Consolidated function to find all changes:
    1. New rows (additions)
    2. Deleted rows (removals)
    3. Cell changes in existing rows
    
    All changes are written to a SINGLE output file with clear categorization.
    Returns: summary statistics
    """
    t0 = time.time()
    if status_callback: status_callback("Building data maps from old and new files...")
    
    # Build complete row data from both files
    old_data, old_header = build_row_data_from_xlsx(old_path, key_cols, sheet_name, progress_callback=progress_callback)
    new_data, new_header = build_row_data_from_xlsx(new_path, key_cols, sheet_name, progress_callback=progress_callback)
    
    if status_callback: 
        status_callback(f"Old file: {len(old_data):,} rows | New file: {len(new_data):,} rows")
        status_callback(f"Key columns used: {', '.join(key_cols)}")

    # Get key column indices
    old_name_to_idx = {str(name).strip(): idx for idx, name in enumerate(old_header)}
    new_name_to_idx = {str(name).strip(): idx for idx, name in enumerate(new_header)}
    
    # Prepare output workbook
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "All Changes"
    
    # Header with formatting
    output_header = ["Change Type", "eNodeB Name", "4LRD"] + new_header
    ws_out.append(output_header)
    
    # Format header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = ExcelFont(bold=True, color="FFFFFF")
    for cell in ws_out[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Track statistics
    enodeb_name_col = None
    for col_name in ["eNodeB Name", "eNodeB_Name", "eNodeBName", "enodeb_name"]:
        if col_name in new_name_to_idx:
            enodeb_name_col = col_name
            break
    
    new_rows_count = 0
    deleted_rows_count = 0
    changed_rows_count = 0
    affected_enodebs = set()
    
    # Color fills for different change types
    new_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    deleted_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    changed_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
    
    if status_callback: status_callback("📝 Processing new rows...")
    
    # 1. Find NEW ROWS (in new_data but not in old_data)
    for key, row_dict in new_data.items():
        if key not in old_data:
            new_rows_count += 1
            enodeb_name = row_dict.get(enodeb_name_col, str(key[0])) if enodeb_name_col else str(key[0])
            affected_enodebs.add(str(enodeb_name))
            
            # Write row
            row_values = [row_dict.get(col, "") for col in new_header]
            four_lrd = str(enodeb_name)[:4] if enodeb_name else ""
            ws_out.append(["NEW ROW", enodeb_name, four_lrd] + row_values)
            
            # Apply green fill
            current_row = ws_out.max_row
            for cell in ws_out[current_row]:
                cell.fill = new_fill
    
    if status_callback: status_callback(f"   Found {new_rows_count:,} new rows")
    if status_callback: status_callback("📝 Processing deleted rows...")
    
    # 2. Find DELETED ROWS (in old_data but not in new_data)
    for key, row_dict in old_data.items():
        if key not in new_data:
            deleted_rows_count += 1
            enodeb_name = row_dict.get(enodeb_name_col, str(key[0])) if enodeb_name_col else str(key[0])
            affected_enodebs.add(str(enodeb_name))
            
            # Write row (use old header structure but map to new header)
            row_values = [row_dict.get(col, "") for col in new_header]
            four_lrd = str(enodeb_name)[:4] if enodeb_name else ""
            ws_out.append(["DELETED ROW", enodeb_name, four_lrd] + row_values)
            
            # Apply red fill
            current_row = ws_out.max_row
            for cell in ws_out[current_row]:
                cell.fill = deleted_fill
    
    if status_callback: status_callback(f"   Found {deleted_rows_count:,} deleted rows")
    if status_callback: status_callback("📝 Processing cell changes in existing rows...")
    
    # 3. Find CELL CHANGES (rows exist in both, but values differ)
    scanned = 0
    for key in new_data.keys():
        if key in old_data:
            scanned += 1
            old_row = old_data[key]
            new_row = new_data[key]
            
            # Check for any cell differences
            has_changes = False
            for col_name in new_header:
                if col_name not in key_cols:  # Don't check key columns
                    old_val = normalize(old_row.get(col_name, ""))
                    new_val = normalize(new_row.get(col_name, ""))
                    if old_val != new_val:
                        has_changes = True
                        break
            
            if has_changes:
                changed_rows_count += 1
                enodeb_name = new_row.get(enodeb_name_col, str(key[0])) if enodeb_name_col else str(key[0])
                affected_enodebs.add(str(enodeb_name))
                
                # Write the NEW values of the changed row
                row_values = [new_row.get(col, "") for col in new_header]
                four_lrd = str(enodeb_name)[:4] if enodeb_name else ""
                ws_out.append(["CELL CHANGE", enodeb_name, four_lrd] + row_values)
                
                # Apply yellow fill
                current_row = ws_out.max_row
                for cell in ws_out[current_row]:
                    cell.fill = changed_fill
            
            if progress_callback and (scanned % 20000 == 0):
                progress_callback(f"Scanned {scanned:,} existing rows for changes...")
    
    if status_callback: status_callback(f"   Found {changed_rows_count:,} rows with cell changes")
    
    # Adjust column widths
    for column in ws_out.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_out.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb_out.save(out_path)
    
    elapsed = time.time() - t0
    if status_callback:
        status_callback(f"✅ All changes written to: {out_path}")
        status_callback(f"📊 Summary:")
        status_callback(f"   • New rows: {new_rows_count:,}")
        status_callback(f"   • Deleted rows: {deleted_rows_count:,}")
        status_callback(f"   • Changed rows: {changed_rows_count:,}")
        status_callback(f"   • Total changes: {new_rows_count + deleted_rows_count + changed_rows_count:,}")
        status_callback(f"   • Affected eNodeBs: {len(affected_enodebs):,}")
        if len(affected_enodebs) <= 30:
            status_callback(f"   • eNodeB Names: {', '.join(sorted(affected_enodebs))}")
        else:
            status_callback(f"   • eNodeB Names (first 30): {', '.join(sorted(list(affected_enodebs)[:30]))}...")
        status_callback(f"⏱️  Processing time: {elapsed:.1f}s")
    
    return {
        'new_rows': new_rows_count,
        'deleted_rows': deleted_rows_count,
        'changed_rows': changed_rows_count,
        'total_changes': new_rows_count + deleted_rows_count + changed_rows_count,
        'affected_enodebs': sorted(affected_enodebs)
    }

def build_key_set_from_csv(p: Path, key_cols: list, progress_callback=None):
    """Build a set of composite keys from CSV file"""
    s = set()
    with p.open("r", encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh)
        header = next(reader)
        name_to_idx = {h.strip(): i for i, h in enumerate(header)}
        
        # Validate all key columns exist
        key_idx = []
        for k in key_cols:
            if k not in name_to_idx:
                raise ValueError(f"Key column '{k}' not found in CSV header. Available: {list(name_to_idx.keys())}")
            key_idx.append(name_to_idx[k])
        
        scanned = 0
        for row in reader:
            scanned += 1
            # Build composite key
            key = tuple(normalize(row[i]) for i in key_idx)
            s.add(key)
            
            if progress_callback and (scanned % 50000 == 0):
                progress_callback(f"Scanned {scanned:,} rows from old CSV...")
    
    return s

def build_row_data_from_csv(p: Path, key_cols: list, progress_callback=None):
    """Build a dictionary mapping keys to full row data from CSV"""
    row_data = {}
    with p.open("r", encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh)
        header = next(reader)
        name_to_idx = {h.strip(): i for i, h in enumerate(header)}
        
        # Validate all key columns exist
        key_idx = []
        for k in key_cols:
            if k not in name_to_idx:
                raise ValueError(f"Key column '{k}' not found in CSV header. Available: {list(name_to_idx.keys())}")
            key_idx.append(name_to_idx[k])
        
        scanned = 0
        for row in reader:
            scanned += 1
            # Build composite key
            key = tuple(normalize(row[i]) for i in key_idx)
            row_data[key] = {header[i]: row[i] for i in range(len(row))}
            
            if progress_callback and (scanned % 50000 == 0):
                progress_callback(f"Scanned {scanned:,} rows from CSV...")
    
    return row_data, header

def find_all_changes_csv_to_xlsx(old_path: Path, new_path: Path, key_cols: list, out_path: Path,
                                 sheet_name=None, status_callback=None, progress_callback=None):
    """
    Consolidated function for CSV old file and Excel new file.
    Finds all changes and writes to single Excel output.
    """
    t0 = time.time()
    if status_callback: status_callback("Building data maps from old CSV and new Excel files...")
    
    # Build complete row data from both files
    old_data, old_header = build_row_data_from_csv(old_path, key_cols, progress_callback=progress_callback)
    new_data, new_header = build_row_data_from_xlsx(new_path, key_cols, sheet_name, progress_callback=progress_callback)
    
    if status_callback: 
        status_callback(f"Old file: {len(old_data):,} rows | New file: {len(new_data):,} rows")
        status_callback(f"Key columns used: {', '.join(key_cols)}")

    # Prepare output workbook
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "All Changes"
    
    # Header with formatting
    output_header = ["Change Type", "eNodeB Name", "4LRD"] + list(new_header)
    ws_out.append(output_header)
    
    # Format header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = ExcelFont(bold=True, color="FFFFFF")
    for cell in ws_out[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Track statistics
    new_name_to_idx = {str(name).strip(): idx for idx, name in enumerate(new_header)}
    enodeb_name_col = None
    for col_name in ["eNodeB Name", "eNodeB_Name", "eNodeBName", "enodeb_name"]:
        if col_name in new_name_to_idx:
            enodeb_name_col = col_name
            break
    
    new_rows_count = 0
    deleted_rows_count = 0
    changed_rows_count = 0
    affected_enodebs = set()
    
    # Color fills
    new_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    deleted_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    changed_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    if status_callback: status_callback("📝 Processing new rows...")
    
    # 1. NEW ROWS
    for key, row_dict in new_data.items():
        if key not in old_data:
            new_rows_count += 1
            enodeb_name = row_dict.get(enodeb_name_col, str(key[0])) if enodeb_name_col else str(key[0])
            affected_enodebs.add(str(enodeb_name))
            
            row_values = [row_dict.get(col, "") for col in new_header]
            four_lrd = str(enodeb_name)[:4] if enodeb_name else ""
            ws_out.append(["NEW ROW", enodeb_name, four_lrd] + row_values)
            
            current_row = ws_out.max_row
            for cell in ws_out[current_row]:
                cell.fill = new_fill
    
    if status_callback: status_callback(f"   Found {new_rows_count:,} new rows")
    if status_callback: status_callback("📝 Processing deleted rows...")
    
    # 2. DELETED ROWS
    for key, row_dict in old_data.items():
        if key not in new_data:
            deleted_rows_count += 1
            enodeb_name = row_dict.get(enodeb_name_col, str(key[0])) if enodeb_name_col else str(key[0])
            affected_enodebs.add(str(enodeb_name))
            
            row_values = [row_dict.get(col, "") for col in new_header]
            four_lrd = str(enodeb_name)[:4] if enodeb_name else ""
            ws_out.append(["DELETED ROW", enodeb_name, four_lrd] + row_values)
            
            current_row = ws_out.max_row
            for cell in ws_out[current_row]:
                cell.fill = deleted_fill
    
    if status_callback: status_callback(f"   Found {deleted_rows_count:,} deleted rows")
    if status_callback: status_callback("📝 Processing cell changes in existing rows...")
    
    # 3. CELL CHANGES
    scanned = 0
    for key in new_data.keys():
        if key in old_data:
            scanned += 1
            old_row = old_data[key]
            new_row = new_data[key]
            
            has_changes = False
            for col_name in new_header:
                if col_name not in key_cols:
                    old_val = normalize(old_row.get(col_name, ""))
                    new_val = normalize(new_row.get(col_name, ""))
                    if old_val != new_val:
                        has_changes = True
                        break
            
            if has_changes:
                changed_rows_count += 1
                enodeb_name = new_row.get(enodeb_name_col, str(key[0])) if enodeb_name_col else str(key[0])
                affected_enodebs.add(str(enodeb_name))
                
                row_values = [new_row.get(col, "") for col in new_header]
                four_lrd = str(enodeb_name)[:4] if enodeb_name else ""
                ws_out.append(["CELL CHANGE", enodeb_name, four_lrd] + row_values)
                
                current_row = ws_out.max_row
                for cell in ws_out[current_row]:
                    cell.fill = changed_fill
            
            if progress_callback and (scanned % 20000 == 0):
                progress_callback(f"Scanned {scanned:,} existing rows for changes...")
    
    if status_callback: status_callback(f"   Found {changed_rows_count:,} rows with cell changes")
    
    # Adjust column widths
    for column in ws_out.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_out.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb_out.save(out_path)
    
    elapsed = time.time() - t0
    if status_callback:
        status_callback(f"✅ All changes written to: {out_path}")
        status_callback(f"📊 Summary:")
        status_callback(f"   • New rows: {new_rows_count:,}")
        status_callback(f"   • Deleted rows: {deleted_rows_count:,}")
        status_callback(f"   • Changed rows: {changed_rows_count:,}")
        status_callback(f"   • Total changes: {new_rows_count + deleted_rows_count + changed_rows_count:,}")
        status_callback(f"   • Affected eNodeBs: {len(affected_enodebs):,}")
        if len(affected_enodebs) <= 30:
            status_callback(f"   • eNodeB Names: {', '.join(sorted(affected_enodebs))}")
        else:
            status_callback(f"   • eNodeB Names (first 30): {', '.join(sorted(list(affected_enodebs)[:30]))}...")
        status_callback(f"⏱️  Processing time: {elapsed:.1f}s")
    
    return {
        'new_rows': new_rows_count,
        'deleted_rows': deleted_rows_count,
        'changed_rows': changed_rows_count,
        'total_changes': new_rows_count + deleted_rows_count + changed_rows_count,
        'affected_enodebs': sorted(affected_enodebs)
    }

# -------------------------
# GUI Application
# -------------------------
class FindNewRowsApp(QMainWindow):
    """Main application window"""
    
    def __init__(self):
        super().__init__()
        self.queue = queue.Queue()
        self.timer = QTimer()
        self.timer.timeout.connect(self.poll_queue)
        self.timer.start(100)
        self.output_file_path = None  # Store the output file path
        self.init_ui()
    
    def init_ui(self):
        """Initialize the user interface"""
        self.setWindowTitle("Track EP Changes - Pro v1.6")
        self.setGeometry(100, 100, 700, 620)
        
        # Set window icon if available
        icon_path = Path(__file__).parent / "app_icon.ico"
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))
        
        # Apply Windows native stylesheet
        self.setStyleSheet(self.get_windows_stylesheet())
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(8)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # Title Section - Compact
        title_layout = QHBoxLayout()
        title_icon = QLabel("📊")
        title_icon.setFont(QFont("Tahoma", 16))
        title_label = QLabel("EP Comparison")
        title_label.setFont(QFont("Tahoma", 14, QFont.Weight.Bold))
        title_label.setStyleSheet("color: #000080;")
        subtitle = QLabel("Track EP changes: new rows, deletions, and cell modifications")
        subtitle.setFont(QFont("Tahoma", 9))
        subtitle.setStyleSheet("color: #666666;")
        
        title_layout.addWidget(title_icon)
        title_layout.addWidget(title_label)
        title_layout.addWidget(subtitle)
        title_layout.addStretch()
        main_layout.addLayout(title_layout)
        
        # Separator line
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        main_layout.addWidget(line)
        
        # File Input Section
        file_group = QGroupBox("📁 Input Files")
        file_group.setFont(QFont("Tahoma", 10, QFont.Weight.Bold))
        file_layout = QVBoxLayout()
        file_layout.setSpacing(8)
        
        # Old file
        old_layout = QHBoxLayout()
        old_layout.setSpacing(8)
        old_label = QLabel("Old File:")
        old_label.setFont(QFont("Tahoma", 9))
        old_label.setFixedWidth(60)
        self.old_file_edit = QLineEdit()
        self.old_file_edit.setFont(QFont("Tahoma", 9))
        self.old_file_edit.setPlaceholderText("Select the old/existing file...")
        old_btn = QPushButton("📂 Browse")
        old_btn.setFont(QFont("Tahoma", 9, QFont.Weight.Bold))
        old_btn.setFixedWidth(90)
        old_btn.clicked.connect(self.browse_old)
        old_layout.addWidget(old_label)
        old_layout.addWidget(self.old_file_edit)
        old_layout.addWidget(old_btn)
        file_layout.addLayout(old_layout)
        
        # New file
        new_layout = QHBoxLayout()
        new_layout.setSpacing(8)
        new_label = QLabel("New File:")
        new_label.setFont(QFont("Tahoma", 9))
        new_label.setFixedWidth(60)
        self.new_file_edit = QLineEdit()
        self.new_file_edit.setFont(QFont("Tahoma", 9))
        self.new_file_edit.setPlaceholderText("Select the new/updated file...")
        self.new_file_edit.textChanged.connect(self.update_default_output)
        new_btn = QPushButton("📂 Browse")
        new_btn.setFont(QFont("Tahoma", 9, QFont.Weight.Bold))
        new_btn.setFixedWidth(90)
        new_btn.clicked.connect(self.browse_new)
        new_layout.addWidget(new_label)
        new_layout.addWidget(self.new_file_edit)
        new_layout.addWidget(new_btn)
        file_layout.addLayout(new_layout)
        
        file_group.setLayout(file_layout)
        main_layout.addWidget(file_group)
        
        # Configuration Section
        config_group = QGroupBox("⚙️ Configuration")
        config_group.setFont(QFont("Tahoma", 10, QFont.Weight.Bold))
        config_layout = QVBoxLayout()
        config_layout.setSpacing(8)
        
        # Key columns
        key_layout = QHBoxLayout()
        key_layout.setSpacing(8)
        key_label = QLabel("Key Columns:")
        key_label.setFont(QFont("Tahoma", 9))
        key_label.setFixedWidth(85)
        key_label.setToolTip("Comma-separated column names to use as unique identifiers")
        self.key_cols_edit = QLineEdit()
        self.key_cols_edit.setFont(QFont("Tahoma", 9))
        self.key_cols_edit.setText("eNodeB Name, CellName")
        self.key_cols_edit.setPlaceholderText("e.g., ID, Email, Phone (comma-separated)")
        key_layout.addWidget(key_label)
        key_layout.addWidget(self.key_cols_edit)
        config_layout.addLayout(key_layout)
        
        # Sheet name
        sheet_layout = QHBoxLayout()
        sheet_layout.setSpacing(8)
        sheet_label = QLabel("Sheet Name:")
        sheet_label.setFont(QFont("Tahoma", 9))
        sheet_label.setFixedWidth(85)
        sheet_label.setToolTip("Optional: specify sheet name for Excel files")
        self.sheet_edit = QLineEdit()
        self.sheet_edit.setFont(QFont("Tahoma", 9))
        self.sheet_edit.setPlaceholderText("Optional (defaults to first sheet)")
        sheet_layout.addWidget(sheet_label)
        sheet_layout.addWidget(self.sheet_edit)
        config_layout.addLayout(sheet_layout)
        
        config_group.setLayout(config_layout)
        main_layout.addWidget(config_group)
        
        # Output Section
        output_group = QGroupBox("💾 Output File")
        output_group.setFont(QFont("Tahoma", 10, QFont.Weight.Bold))
        output_layout = QVBoxLayout()
        output_layout.setSpacing(8)
        
        # Output file path (read-only, auto-generated)
        out_path_layout = QHBoxLayout()
        out_path_layout.setSpacing(8)
        out_label = QLabel("Save As:")
        out_label.setFont(QFont("Tahoma", 9))
        out_label.setFixedWidth(60)
        self.out_file_edit = QLineEdit()
        self.out_file_edit.setFont(QFont("Tahoma", 9))
        self.out_file_edit.setPlaceholderText("Auto-generated: Changes_Report_[timestamp].xlsx in source folder")
        self.out_file_edit.setReadOnly(True)
        self.out_file_edit.setStyleSheet("QLineEdit { background-color: #F0F0F0; color: #666666; }")
        out_path_layout.addWidget(out_label)
        out_path_layout.addWidget(self.out_file_edit)
        output_layout.addLayout(out_path_layout)
        
        output_group.setLayout(output_layout)
        main_layout.addWidget(output_group)
        
        # Progress Bar
        progress_layout = QHBoxLayout()
        progress_icon = QLabel("⏳")
        progress_icon.setFont(QFont("Tahoma", 11))
        self.progress_bar = QProgressBar()
        self.progress_bar.setMaximumHeight(18)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setFont(QFont("Tahoma", 9))
        progress_layout.addWidget(progress_icon)
        progress_layout.addWidget(self.progress_bar)
        main_layout.addLayout(progress_layout)
        
        # Action Buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        
        self.start_btn = QPushButton("▶️ Start Processing")
        self.start_btn.setFont(QFont("Tahoma", 11, QFont.Weight.Bold))
        self.start_btn.setMinimumHeight(32)
        self.start_btn.clicked.connect(self.start_process)
        
        self.open_folder_btn = QPushButton("📂 Output Folder")
        self.open_folder_btn.setFont(QFont("Tahoma", 10, QFont.Weight.Bold))
        self.open_folder_btn.setMinimumHeight(32)
        self.open_folder_btn.setMaximumWidth(140)
        self.open_folder_btn.clicked.connect(self.open_output_folder)
        self.open_folder_btn.setEnabled(False)
        
        clear_btn = QPushButton("🗑️ Clear Log")
        clear_btn.setFont(QFont("Tahoma", 10, QFont.Weight.Bold))
        clear_btn.setMinimumHeight(32)
        clear_btn.setMaximumWidth(120)
        clear_btn.clicked.connect(self.clear_log)
        
        button_layout.addWidget(self.start_btn)
        button_layout.addWidget(self.open_folder_btn)
        button_layout.addWidget(clear_btn)
        main_layout.addLayout(button_layout)
        
        # Status Log Section
        status_header_layout = QHBoxLayout()
        status_icon = QLabel("📝")
        status_icon.setFont(QFont("Tahoma", 11))
        status_label = QLabel("Process Log:")
        status_label.setFont(QFont("Tahoma", 10, QFont.Weight.Bold))
        status_label.setStyleSheet("font-weight: bold; color: #000000;")
        status_header_layout.addWidget(status_icon)
        status_header_layout.addWidget(status_label)
        status_header_layout.addStretch()
        main_layout.addLayout(status_header_layout)
        
        self.status_text = QTextEdit()
        self.status_text.setReadOnly(True)
        self.status_text.setMaximumHeight(160)
        self.status_text.setFont(QFont("Tahoma", 9))
        self.status_text.setPlaceholderText("Process logs will appear here...")
        main_layout.addWidget(self.status_text)
        
        # Footer
        footer_label = QLabel("V1.6.4426 • Written in Python. By Fadzli Abdullah ❤️")
        footer_label.setFont(QFont("Tahoma", 8))
        footer_label.setStyleSheet("color: #666666; margin-top: 4px;")
        footer_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(footer_label)

    def get_windows_stylesheet(self):
        """Windows native theme stylesheet"""
        return """
            QMainWindow {
                background-color: #F0F0F0;
            }
            QWidget {
                background-color: #F0F0F0;
                font-family: 'Tahoma', 'Tahoma', sans-serif;
            }
            QGroupBox {
                background-color: #F0F0F0;
                border: 2px groove #A0A0A0;
                border-radius: 0px;
                margin-top: 10px;
                padding: 4px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 2px 6px;
                color: #000080;
                background-color: #F0F0F0;
            }
            QLineEdit {
                padding: 4px 6px;
                border: 2px inset #808080;
                border-radius: 0px;
                background-color: #FFFFFF;
                min-height: 20px;
                color: #000000;
                selection-background-color: #0078D7;
                selection-color: #FFFFFF;
            }
            QLineEdit:focus {
                border: 2px inset #0078D7;
                background-color: #FFFFFF;
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #FFFFFF, stop:1 #E0E0E0);
                color: #000000;
                border: 1px solid #A0A0A0;
                border-radius: 0px;
                padding: 5px 12px;
                font-weight: bold;
                min-height: 24px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #E5F3FF, stop:1 #CCE8FF);
                border: 1px solid #0078D7;
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 #CCE8FF, stop:1 #B3DEFF);
                border: 1px solid #005499;
                padding-left: 6px;
                padding-top: 6px;
            }
            QPushButton:disabled {
                background: #D0D0D0;
                color: #808080;
                border: 1px solid #A0A0A0;
            }
            QTextEdit {
                border: 2px inset #808080;
                border-radius: 0px;
                background-color: #FFFFFF;
                padding: 6px;
                color: #000000;
                selection-background-color: #0078D7;
                selection-color: #FFFFFF;
            }
            QProgressBar {
                border: 2px inset #808080;
                border-radius: 0px;
                background-color: #FFFFFF;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #0078D7;
                border-radius: 0px;
            }
            QCheckBox {
                spacing: 4px;
                color: #000000;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
                border: 2px inset #808080;
                background-color: #FFFFFF;
            }
            QCheckBox::indicator:checked {
                background-color: #FFFFFF;
                border: 2px inset #808080;
                image: none;
            }
            QLabel {
                color: #000000;
                background-color: transparent;
            }
            QFrame[frameShape="4"] {
                color: #A0A0A0;
            }
        """

    def clear_log(self):
        """Clear the status log"""
        self.status_text.clear()
    
    def update_default_output(self):
        """Update the default output filename based on new file path"""
        new_file = self.new_file_edit.text().strip()
        if new_file:
            new_path = Path(new_file)
            if new_path.exists():
                # Generate timestamp
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                
                # Create output filename with timestamp - always Excel
                output_filename = f"Changes_Report_{timestamp}.xlsx"
                output_path = new_path.parent / output_filename
                
                self.out_file_edit.setText(str(output_path))
                self.output_file_path = output_path

    def browse_old(self):
        """Browse for old file"""
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Old File",
            "",
            "Excel and CSV Files (*.xlsx *.xlsm *.csv);;Excel Files (*.xlsx *.xlsm);;CSV Files (*.csv)"
        )
        if path:
            self.old_file_edit.setText(path)

    def browse_new(self):
        """Browse for new file"""
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select New File",
            "",
            "Excel and CSV Files (*.xlsx *.xlsm *.csv);;Excel Files (*.xlsx *.xlsm);;CSV Files (*.csv)"
        )
        if path:
            self.new_file_edit.setText(path)
    
    def open_output_folder(self):
        """Open the folder containing the output file"""
        if self.output_file_path and self.output_file_path.exists():
            folder_path = str(self.output_file_path.parent)
            
            # Open folder based on operating system
            if platform.system() == "Windows":
                os.startfile(folder_path)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", folder_path])
            else:  # Linux and other Unix-like systems
                subprocess.run(["xdg-open", folder_path])
        else:
            QMessageBox.warning(
                self,
                "Folder Not Found",
                "Output file does not exist yet. Please run the process first."
            )

    def append_status(self, msg: str):
        """Append message to status text area"""
        self.status_text.append(msg)
        # Auto-scroll to bottom
        cursor = self.status_text.textCursor()
        cursor.movePosition(cursor.MoveOperation.End)
        self.status_text.setTextCursor(cursor)
    
    def start_process(self):
        """Start the comparison process in a background thread"""
        old = self.old_file_edit.text().strip()
        new = self.new_file_edit.text().strip()
        keys = self.key_cols_edit.text().strip()
        sheet = self.sheet_edit.text().strip() or None
        out = self.out_file_edit.text().strip()

        # Validation
        if not old or not new or not keys:
            QMessageBox.warning(
                self,
                "Missing Information",
                "Please select old file, new file, and key column(s)."
            )
            return
        
        if not out:
            QMessageBox.warning(
                self,
                "Missing Output",
                "Output file path was not generated. Please select a new file first."
            )
            return

        old_path = Path(old)
        new_path = Path(new)
        out_path = Path(out)
        
        # Ensure output is Excel
        if not out_path.suffix:
            out_path = out_path.with_suffix(".xlsx")

        if not old_path.exists() or not new_path.exists():
            QMessageBox.critical(
                self,
                "File Not Found",
                "Old or New file path does not exist."
            )
            return

        # Store output path for folder opening
        self.output_file_path = out_path
        
        # Disable UI
        self.start_btn.setEnabled(False)
        self.open_folder_btn.setEnabled(False)
        self.progress_bar.setRange(0, 0)  # Indeterminate mode
        self.append_status("🚀 Starting process...")
        self.append_status(f"Using key columns: {keys}")

        # Run processing in background thread
        key_cols_list = [k.strip() for k in keys.split(",") if k.strip()]
        thread = threading.Thread(
            target=self._worker_thread,
            args=(old_path, new_path, key_cols_list, sheet, out_path),
            daemon=True
        )
        thread.start()

    def _worker_thread(self, old_path, new_path, key_cols, sheet_name, out_path):
        """Worker thread for processing"""
        try:
            def status_cb(msg):
                self.queue.put(("status", msg))

            def progress_cb(msg):
                self.queue.put(("progress", msg))

            old_ext = old_path.suffix.lower()
            new_ext = new_path.suffix.lower()
            start_time = time.time()
            
            if old_ext in [".xlsx", ".xlsm", ".xltx", ".xltm"] and new_ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
                # Both Excel files
                status_cb("📊 Processing Excel files...")
                result = find_all_changes_xlsx(
                    old_path, new_path, key_cols, out_path,
                    sheet_name=sheet_name, out_xlsx=True,
                    status_callback=status_cb, progress_callback=progress_cb
                )
                    
            elif old_ext == ".csv" and new_ext in [".xlsx", ".xlsm"]:
                # CSV old, Excel new
                status_cb("📊 Processing CSV to Excel comparison...")
                result = find_all_changes_csv_to_xlsx(
                    old_path, new_path, key_cols, out_path,
                    sheet_name=sheet_name,
                    status_callback=status_cb, progress_callback=progress_cb
                )
                        
            else:
                status_cb("⚠ Unsupported file type combination.")
                status_cb("ℹ️  Supported: Both Excel, or CSV (old) + Excel (new)")
                raise ValueError("Unsupported combination of file types.")
            
            elapsed = time.time() - start_time
            self.queue.put(("done", None))
            
        except Exception as e:
            tb = traceback.format_exc()
            self.queue.put(("error", f"{str(e)}\n{tb}"))

    def poll_queue(self):
        """Poll the queue for messages from worker thread"""
        try:
            while True:
                item = self.queue.get_nowait()
                typ, payload = item
                
                if typ == "status":
                    self.append_status(payload)
                elif typ == "progress":
                    self.append_status(f"  ⏳ {payload}")
                elif typ == "error":
                    self.append_status(f"❌ ERROR: {str(payload)}")
                    QMessageBox.critical(
                        self,
                        "Error",
                        str(payload).splitlines()[0]
                    )
                    self._finish_run()
                elif typ == "done":
                    self.append_status("✨ Process finished successfully!")
                    self._finish_run()
                    # Enable the output folder button after successful completion
                    self.open_folder_btn.setEnabled(True)
        except queue.Empty:
            pass

    def _finish_run(self):
        """Clean up after process finishes"""
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(100)
        self.start_btn.setEnabled(True)


def main():
    """Main entry point"""
    app = QApplication(sys.argv)
    app.setStyle('Fusion')  # Use Fusion style for better cross-platform consistency
    
    window = FindNewRowsApp()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
